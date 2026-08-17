"""Mud Report – bitácora de propiedades de fluidos."""
import hashlib
import io
import os
import re
import textwrap
import time
import smtplib
from datetime import datetime
from email.message import EmailMessage

import numpy as np
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st

# Versión de la app, junto al título. SUBIR EN CADA CAMBIO que se despliegue: el número
# sube y la fecha es la del cambio. Van en una sola constante a propósito, para que no
# puedan quedar desincronizadas, y APP_VERSION_NOTA describe qué trae esta versión (se
# muestra al pasar el cursor por encima).
APP_VERSION = "1.4 · 2026-08-17"
APP_VERSION_NOTA = (
    "Reporte Mi SWACO en Excel, encabezados con las etiquetas del propio reporte "
    "(selector Español/Inglés) y solo las columnas que el reporte trae."
)

PLOTLY_CONFIG = {"displayModeBar": False, "displaylogo": False}
PLOTLY_TEMPLATE = "plotly_white"
MUD_SRC_FILES = "files"
MUD_SRC_EMAIL = "email"


def _mud_secret(name: str, default=""):
    env_val = os.getenv(name)
    if env_val is not None and str(env_val).strip():
        return str(env_val).strip()
    try:
        if name in st.secrets:
            return str(st.secrets[name]).strip()
    except Exception:
        pass
    return default


MUD_SMTP_SERVER = _mud_secret("MUD_SMTP_SERVER", _mud_secret("SMTP_SERVER", "smtp.gmail.com"))
MUD_SMTP_PORT = int(_mud_secret("MUD_SMTP_PORT", _mud_secret("SMTP_PORT", "587")))
MUD_SMTP_USER = _mud_secret("MUD_SMTP_USER", _mud_secret("SMTP_USER", ""))
MUD_SMTP_PASS = _mud_secret("MUD_SMTP_PASS", _mud_secret("SMTP_PASS", ""))
MUD_SMTP_FROM = _mud_secret("MUD_SMTP_FROM", MUD_SMTP_USER)
MUD_SMTP_TO = _mud_secret("MUD_SMTP_TO", _mud_secret("SMTP_TO", _mud_secret("TO_EMAIL", "solobox+pemex@rogii.com")))

MUD_IMAP_SERVER = _mud_secret("MUD_IMAP_SERVER", _mud_secret("IMAP_SERVER", "imap.gmail.com"))
MUD_IMAP_USER = _mud_secret("MUD_IMAP_USER", _mud_secret("IMAP_USER", MUD_SMTP_USER))
MUD_IMAP_PASS = _mud_secret("MUD_IMAP_PASS", _mud_secret("IMAP_PASS", MUD_SMTP_PASS))
MUD_IMAP_FILTER = _mud_secret("MUD_IMAP_FILTER", "")


def _safe_numeric_series(df: pd.DataFrame, col: str) -> pd.Series:
    if col not in df.columns:
        return pd.Series(dtype=float)
    vals = df[col]
    if isinstance(vals, pd.DataFrame):
        best = vals.iloc[:, 0]
        best_n = pd.to_numeric(best, errors="coerce").notna().sum()
        for i in range(1, vals.shape[1]):
            cand = vals.iloc[:, i]
            cand_n = pd.to_numeric(cand, errors="coerce").notna().sum()
            if cand_n > best_n:
                best, best_n = cand, cand_n
        vals = best
    return pd.to_numeric(vals, errors="coerce")


def is_streamlit_dark_mode() -> bool:
    try:
        return str(st.get_option("theme.base")).lower() == "dark"
    except Exception:
        return False


def apply_pro_theme(fig, h: int = 420):
    fig.update_layout(
        template=PLOTLY_TEMPLATE,
        height=h,
        margin=dict(l=50, r=30, t=40, b=55),
        title=dict(x=0.02, xanchor="left"),
        font=dict(family="Segoe UI", size=12, color="#2A2A2A"),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
    )
    fig.update_xaxes(showgrid=True, gridcolor="rgba(0,0,0,0.06)", zeroline=False)
    fig.update_yaxes(showgrid=True, gridcolor="rgba(0,0,0,0.06)", zeroline=False)
    return fig


def apply_pro_theme_dark(fig, h: int = 420):
    fig.update_layout(
        template="plotly_dark",
        height=h,
        margin=dict(l=50, r=30, t=40, b=55),
        title=dict(x=0.02, xanchor="left"),
        font=dict(family="Segoe UI", size=12, color="#E5E7EB"),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
    )
    fig.update_xaxes(showgrid=True, gridcolor="rgba(255,255,255,0.08)", zeroline=False)
    fig.update_yaxes(showgrid=True, gridcolor="rgba(255,255,255,0.08)", zeroline=False)
    return fig


def prettify_auto(fig, h: int = 420):
    return apply_pro_theme_dark(fig, h=h) if is_streamlit_dark_mode() else apply_pro_theme(fig, h=h)


def prettify(fig, h: int = 420):
    return apply_pro_theme(fig, h=h)


def prettify_hist(fig, h: int = 420):
    fig.update_layout(template=PLOTLY_TEMPLATE, height=h, margin=dict(l=50, r=30, t=18, b=55))
    return fig


def prettify_heatmap(fig, h: int = 520):
    fig.update_layout(template=PLOTLY_TEMPLATE, height=h, margin=dict(l=60, r=30, t=48, b=60))
    return fig


def prettify_heatmap_auto(fig, h: int = 520):
    if is_streamlit_dark_mode():
        fig.update_layout(template="plotly_dark", height=h, paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(15,23,42,0.5)")
        return fig
    return prettify_heatmap(fig, h=h)


def apply_line_area_fill(fig, line_color: str | None = None, fill_alpha: float = 0.22, line_width: float = 2.0, skip_dashed: bool = False) -> go.Figure:
    from plotly.colors import hex_to_rgb

    def _rgba(color, alpha: float) -> str:
        if not color or not str(color).startswith("#"):
            return f"rgba(37,99,235,{alpha})"
        try:
            t = hex_to_rgb(color)
            return f"rgba({int(t[0])},{int(t[1])},{int(t[2])},{alpha})"
        except Exception:
            return f"rgba(37,99,235,{alpha})"

    palette = ["#2563EB", "#EA580C", "#10B981", "#8B5CF6"]
    for i, trace in enumerate(fig.data):
        if getattr(trace, "type", None) != "scatter" or "lines" not in (trace.mode or "lines"):
            continue
        lc = getattr(getattr(trace, "line", None), "color", None) or line_color or palette[i % len(palette)]
        try:
            fig.data[i].update(fill="tozeroy", fillcolor=_rgba(lc, fill_alpha), line=dict(width=line_width, color=lc))
        except Exception:
            pass
    return fig


# =========================
# Sistema visual del Mud Report (gráficas HD)
# =========================
# Paleta categórica validada (banda de luminosidad, piso de croma, separación CVD
# adyacente y contraste) contra las superficies reales de Streamlit: #FFFFFF en claro
# y #0E1117 en oscuro. La columna oscura son los mismos ocho tonos re-escalonados
# para fondo oscuro, no un volteo automático. El ORDEN es el mecanismo de seguridad
# CVD: no reordenar ni generar tonos extra más allá del octavo.
_MUD_CAT_LIGHT = ["#2a78d6", "#eb6834", "#1baf7a", "#eda100", "#e87ba4", "#008300", "#4a3aa7", "#e34948"]
_MUD_CAT_DARK = ["#3987e5", "#d95926", "#199e70", "#c98500", "#d55181", "#008300", "#9085e9", "#e66767"]
MUD_MAX_OVERLAY_SERIES = len(_MUD_CAT_LIGHT)
_MUD_MAX_PANELS = 12
# Propiedades que un ingeniero de lodos sigue jornada a jornada.
MUD_DEFAULT_CHART_PROPS = ["Density", "PV", "YP", "FV"]
_MUD_EVO_PANELS = "panels"
_MUD_EVO_NORM = "norm"
_MUD_EVO_RAW = "raw"

# Estados: fijos, nunca tematizados, nunca reutilizados como color de serie.
_MUD_STATUS = {"good": "#0ca30c", "warning": "#fab219", "serious": "#ec835a", "critical": "#d03b3b"}

_MUD_TOKENS_LIGHT = {
    "ink": "#0b0b0b", "secondary": "#52514e", "muted": "#898781",
    "grid": "#e1e0d9", "axis": "#c3c2b7", "surface": "#FFFFFF",
    "hover_bg": "rgba(255,255,255,0.97)", "ring": "#FFFFFF", "neutral": "#f0efec",
}
_MUD_TOKENS_DARK = {
    "ink": "#ffffff", "secondary": "#c3c2b7", "muted": "#898781",
    "grid": "#2c2c2a", "axis": "#383835", "surface": "#0E1117",
    "hover_bg": "rgba(13,13,13,0.95)", "ring": "#0E1117", "neutral": "#383835",
}

# Fechas en formato numérico (%d/%m), no %b: plotly.js rotula los meses en inglés salvo
# que se cargue el bundle de idioma, que Streamlit no incluye.
_MUD_TIME_TICKS = [
    dict(dtickrange=[None, 3_600_000], value="%H:%M"),
    dict(dtickrange=[3_600_000, 43_200_000], value="%d/%m %H:%M"),
    dict(dtickrange=[43_200_000, None], value="%d/%m"),
]


def _mud_viz_tokens() -> dict:
    return _MUD_TOKENS_DARK if is_streamlit_dark_mode() else _MUD_TOKENS_LIGHT


def _mud_palette() -> list[str]:
    return _MUD_CAT_DARK if is_streamlit_dark_mode() else _MUD_CAT_LIGHT


def _mud_rgba(color: str, alpha: float) -> str:
    from plotly.colors import hex_to_rgb

    try:
        r, g, b = hex_to_rgb(color)
        return f"rgba({int(r)},{int(g)},{int(b)},{alpha})"
    except Exception:
        return f"rgba(42,120,214,{alpha})"


def _mud_assign_series_slots(selected: list[str]) -> dict[str, int]:
    """
    Slot de color estable por propiedad. Al destildar una propiedad, las que siguen
    en pantalla conservan su color (el color sigue a la entidad, no a su posición en
    la lista): de otro modo quien aprendió «Densidad = azul» queda desorientado.
    """
    store = st.session_state.setdefault("mud_color_slots", {})
    n = len(_MUD_CAT_LIGHT)
    out: dict[str, int] = {}
    used: set[int] = set()
    for p in selected:
        s = store.get(p)
        if s is not None and s not in used:
            out[p] = s
            used.add(s)
    for p in selected:
        if p in out:
            continue
        s = next((k for k in range(n) if k not in used), len(used) % n)
        out[p] = s
        used.add(s)
    store.update(out)
    return out


def _mud_hd_config(name: str, width: int = 1800, height: int = 1000) -> dict:
    """
    Config de Plotly con exportación en alta definición: el botón de descarga entrega
    un PNG a escala 3 (1800x1000 -> 5400x3000 px), suficiente para informe impreso o
    presentación. Es exportación del lado del navegador, sin dependencias nuevas.
    """
    return {
        "displaylogo": False,
        "displayModeBar": True,
        "modeBarButtonsToRemove": ["lasso2d", "select2d", "autoScale2d", "toggleSpikelines"],
        "toImageButtonOptions": {
            "format": "png",
            "filename": _sanitize_filename(name, "mud_chart"),
            "scale": 3,
            "width": width,
            "height": height,
        },
        "responsive": True,
        "scrollZoom": False,
    }


def _mud_hd_theme(fig, h: int = 460, *, title: str = "", legend: bool = True,
                  hovermode: str = "x unified", spikes: bool = True, y_title: str = "") -> go.Figure:
    """Tema único de las gráficas del Mud Report: tipografía y rejilla recesivas,
    hairlines sólidas (nunca punteadas), fondo transparente y hover de cruceta."""
    t = _mud_viz_tokens()
    fig.update_layout(
        template="plotly_dark" if is_streamlit_dark_mode() else "plotly_white",
        height=h,
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        colorway=_mud_palette(),
        font=dict(family='system-ui, -apple-system, "Segoe UI", sans-serif', size=13, color=t["ink"]),
        title=dict(text=title or None, x=0.0, xanchor="left", y=0.98, yanchor="top",
                   font=dict(size=17, color=t["ink"])),
        margin=dict(l=68, r=26, t=(74 if (title and legend) else 52 if (title or legend) else 24), b=54),
        hovermode=hovermode,
        hoverlabel=dict(
            font=dict(family='system-ui, -apple-system, "Segoe UI", sans-serif', size=12),
            bgcolor=t["hover_bg"], bordercolor=t["axis"], namelength=-1,
        ),
        showlegend=legend,
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="left", x=0,
                    font=dict(size=12), bgcolor="rgba(0,0,0,0)", itemsizing="constant",
                    itemwidth=30),
        dragmode="pan",
    )
    fig.update_xaxes(
        showgrid=True, gridcolor=t["grid"], gridwidth=1, zeroline=False,
        showline=True, linecolor=t["axis"], linewidth=1,
        ticks="outside", ticklen=5, tickcolor=t["axis"],
        tickfont=dict(size=11, color=t["muted"]), title_font=dict(size=12, color=t["secondary"]),
        showspikes=spikes, spikemode="across", spikethickness=1, spikecolor=t["muted"], spikedash="dot",
    )
    fig.update_yaxes(
        showgrid=True, gridcolor=t["grid"], gridwidth=1, zeroline=False, showline=False,
        ticks="outside", ticklen=5, tickcolor=t["axis"],
        tickfont=dict(size=11, color=t["muted"]), title_font=dict(size=12, color=t["secondary"]),
    )
    if y_title:
        fig.update_yaxes(title_text=y_title)
    return fig


def _mud_style_time_axis(fig, dates, *, row: str | int | None = None) -> None:
    """Ticks de tiempo adaptativos y una hairline sólida en cada cambio de día, que es
    lo que vuelve legible una bitácora acumulada de varias jornadas."""
    t = _mud_viz_tokens()
    fig.update_xaxes(tickformatstops=_MUD_TIME_TICKS)
    days = pd.to_datetime(pd.Series(list(dates)), errors="coerce").dropna().dt.normalize().unique()
    if not (2 <= len(days) <= 31):
        return
    for d in sorted(days)[1:]:
        fig.add_shape(
            type="line", xref="x", yref="paper",
            x0=pd.Timestamp(d), x1=pd.Timestamp(d), y0=0, y1=1,
            line=dict(color=t["grid"], width=1), layer="below",
        )


def _mud_normalize_0_100(s: pd.Series) -> pd.Series:
    lo, hi = float(s.min()), float(s.max())
    if not np.isfinite(lo) or not np.isfinite(hi) or hi == lo:
        return pd.Series(50.0, index=s.index)
    return (s - lo) / (hi - lo) * 100.0


def _mud_evolution_panels_figure(df: pd.DataFrame, props: list[str]) -> go.Figure:
    """
    Small multiples: un panel por propiedad con eje X compartido y eje Y propio. Es la
    forma correcta cuando las magnitudes no son comparables (Densidad ~1680 kg/m³ vs
    YP ~10 lb/100ft²): en un solo eje todo lo que no sea la densidad queda plano.
    Un único color, porque la identidad la lleva el título de cada eje Y.
    """
    from plotly.subplots import make_subplots

    t = _mud_viz_tokens()
    color = _mud_palette()[0]
    n = len(props)
    fig = make_subplots(rows=n, cols=1, shared_xaxes=True,
                        vertical_spacing=min(0.055, 0.75 / max(1, n - 1)) if n > 1 else 0.0)
    for i, p in enumerate(props, start=1):
        s = df[["Date", p]].dropna().sort_values("Date")
        fig.add_trace(
            go.Scatter(
                x=s["Date"], y=s[p], mode="lines+markers", name=p,
                line=dict(width=2, color=color, shape="linear"),
                marker=dict(size=8, color=color, line=dict(width=2, color=t["ring"])),
                hovertemplate=f"<b>{p}</b> %{{y}}<extra></extra>",
                showlegend=False,
            ),
            row=i, col=1,
        )
        fig.update_yaxes(title_text=p, row=i, col=1, nticks=5)
    fig = _mud_hd_theme(fig, h=90 + 150 * n, legend=False, title="")
    fig.update_layout(margin=dict(l=84, r=26, t=26, b=54))
    fig.update_xaxes(title_text="Fecha / hora", row=n, col=1)
    _mud_style_time_axis(fig, df["Date"])
    return fig


def _mud_evolution_overlay_figure(df: pd.DataFrame, props: list[str], slots: dict[str, int],
                                  *, normalize: bool) -> go.Figure:
    """
    Series superpuestas en UN solo eje. Con normalize=True cada serie va indexada 0–100
    sobre su propio min–max, que es la manera honesta de comparar formas entre
    magnitudes distintas (nunca un segundo eje Y: la alineación entre dos escalas es
    arbitraria e inventa correlaciones). El tooltip siempre muestra el valor real.
    """
    t = _mud_viz_tokens()
    pal = _mud_palette()
    fig = go.Figure()
    for p in props:
        s = df[["Date", p]].dropna().sort_values("Date")
        if s.empty:
            continue
        color = pal[slots.get(p, 0) % len(pal)]
        y = _mud_normalize_0_100(s[p]) if normalize else s[p]
        fig.add_trace(
            go.Scatter(
                x=s["Date"], y=y, mode="lines+markers", name=p,
                line=dict(width=2, color=color),
                marker=dict(size=8, color=color, line=dict(width=2, color=t["ring"])),
                customdata=s[p],
                hovertemplate=f"<b>{p}</b> %{{customdata}}<extra></extra>",
            )
        )
    # Etiqueta directa en el extremo con ≤4 series: así la identidad no depende solo del
    # color (tres tonos claros de la paleta quedan bajo 3:1 contra fondo blanco).
    if 1 <= len(fig.data) <= 4:
        ends = []
        for tr in fig.data:
            ys = [v for v in tr.y if v is not None and np.isfinite(v)]
            if not len(tr.x) or not ys:
                continue
            ends.append([float(ys[-1]), pd.Timestamp(tr.x[-1]).to_pydatetime(), tr.name, tr.line.color])
        all_y = [float(v) for tr in fig.data for v in tr.y if v is not None and np.isfinite(v)]
        if all_y:
            span = max(all_y) - min(all_y)
            gap = (span if span else abs(max(all_y)) or 1.0) * 0.065
        else:
            gap = 1.0
        # Series que terminan en el mismo valor (habitual en normalizado) pisarían su
        # etiqueta: se separan verticalmente lo mínimo para que ambas se lean.
        ends.sort(key=lambda e: e[0], reverse=True)
        for i in range(1, len(ends)):
            if ends[i - 1][0] - ends[i][0] < gap:
                ends[i][0] = ends[i - 1][0] - gap
        for y_lbl, x_lbl, name, color in ends:
            fig.add_annotation(
                x=x_lbl, y=y_lbl, text=f" {name}", showarrow=False,
                xanchor="left", yanchor="middle", font=dict(size=11, color=color),
            )
        labelled = bool(ends)
    else:
        labelled = False
    fig = _mud_hd_theme(
        fig,
        h=480,
        legend=len(fig.data) > 1,
        y_title="Índice 0–100 (min–max de cada propiedad)" if normalize else "Valor",
    )
    if labelled:
        # Después del tema: _mud_hd_theme reescribe el margen completo.
        fig.update_layout(margin=dict(l=68, r=124, t=52, b=54))
    fig.update_xaxes(title_text="Fecha / hora")
    _mud_style_time_axis(fig, df["Date"])
    return fig


def _mud_property_detail_figure(df: pd.DataFrame, prop: str, color: str) -> go.Figure:
    """Serie única con banda media ±1σ, para ver de un golpe qué tan estable estuvo."""
    t = _mud_viz_tokens()
    s = df[["Date", prop]].dropna().sort_values("Date")
    fig = go.Figure()
    if len(s) >= 2:
        mean_v = float(s[prop].mean())
        std_v = float(s[prop].std() or 0.0)
        fig.add_trace(go.Scatter(
            x=list(s["Date"]) + list(s["Date"])[::-1],
            y=[mean_v + std_v] * len(s) + [mean_v - std_v] * len(s),
            fill="toself", fillcolor=_mud_rgba(color, 0.10),
            line=dict(width=0), hoverinfo="skip", name="Media ±1σ",
        ))
        fig.add_hline(y=mean_v, line=dict(color=t["secondary"], width=1, dash="dash"),
                      annotation_text=f"Media {format_num(mean_v)}",
                      annotation_position="top left",
                      annotation_font=dict(size=11, color=t["secondary"]))
    fig.add_trace(go.Scatter(
        x=s["Date"], y=s[prop], mode="lines+markers", name=prop,
        line=dict(width=2, color=color),
        marker=dict(size=9, color=color, line=dict(width=2, color=t["ring"])),
        hovertemplate=f"<b>{prop}</b> %{{y}}<extra></extra>",
    ))
    fig = _mud_hd_theme(fig, h=430, title=f"Evolución — {prop}", y_title=prop)
    fig.update_xaxes(title_text="Fecha / hora")
    _mud_style_time_axis(fig, s["Date"])
    return fig


def _mud_distribution_figure(values, prop: str, color: str, nbins: int = 25) -> go.Figure:
    """Histograma con media y mediana marcadas (línea discontinua = referencia, no rejilla)."""
    t = _mud_viz_tokens()
    vals = pd.Series(values).dropna()
    fig = go.Figure()
    if vals.empty:
        return _mud_hd_theme(fig, h=380, title=f"Distribución — {prop}", legend=False)
    fig.add_trace(go.Histogram(
        x=vals, nbinsx=nbins, name=prop,
        marker=dict(color=_mud_rgba(color, 0.75), line=dict(width=1, color=t["ring"])),
        hovertemplate="%{x}: %{y} muestra(s)<extra></extra>",
    ))
    for label, val, dash in (("Media", float(vals.mean()), "dash"), ("Mediana", float(vals.median()), "dot")):
        fig.add_vline(x=val, line=dict(color=t["secondary"], width=1, dash=dash),
                      annotation_text=f"{label} {format_num(val)}", annotation_position="top",
                      annotation_font=dict(size=11, color=t["secondary"]))
    fig = _mud_hd_theme(fig, h=400, title=f"Distribución — {prop}", legend=False,
                        hovermode="closest", spikes=False, y_title="Muestras")
    fig.update_xaxes(title_text=prop)
    fig.update_layout(bargap=0.06)
    return fig


def _mud_variability_figure(df: pd.DataFrame, props: list[str], top_labels: int = 5) -> go.Figure:
    """
    Ranking de estabilidad por coeficiente de variación (σ/|μ| en %), ordenado de mayor a
    menor. El CV es lo que permite comparar cuánto se movió cada propiedad entre unidades
    distintas — normalizar min–max no serviría: dejaría a todas ocupando exactamente el
    mismo rango y borraría la comparación.
    """
    t = _mud_viz_tokens()
    color = _mud_palette()[0]
    rows = []
    for p in props:
        s = pd.to_numeric(df[p], errors="coerce").dropna()
        if len(s) < 2:
            continue
        mean_v = float(s.mean())
        if not mean_v:
            continue
        rows.append((p, float(s.std()) / abs(mean_v) * 100.0, float(s.min()), float(s.max())))
    fig = go.Figure()
    if not rows:
        return _mud_hd_theme(fig, h=320, legend=False, hovermode="closest", spikes=False)

    rows.sort(key=lambda r: r[1])
    names = [r[0] for r in rows]
    cvs = [r[1] for r in rows]
    fig.add_trace(go.Bar(
        x=cvs, y=names, orientation="h",
        marker=dict(color=color, line=dict(width=0)),
        width=0.62,
        customdata=[[r[2], r[3]] for r in rows],
        hovertemplate="<b>%{y}</b><br>CV %{x:.1f} %<br>rango %{customdata[0]} – %{customdata[1]}<extra></extra>",
        showlegend=False,
    ))
    # Etiqueta solo en las más inestables: el eje y el tooltip cargan el resto.
    for name, cv, _lo, _hi in rows[-top_labels:]:
        fig.add_annotation(x=cv, y=name, text=f" {cv:.1f} %", showarrow=False,
                           xanchor="left", yanchor="middle",
                           font=dict(size=11, color=t["secondary"]))
    fig = _mud_hd_theme(fig, h=max(320, 26 * len(rows) + 96), legend=False,
                        hovermode="closest", spikes=False)
    fig.update_xaxes(title_text="Coeficiente de variación (σ / media, %)",
                     range=[0, max(cvs) * 1.18 if max(cvs) else 1])
    fig.update_yaxes(ticks="", showgrid=False)
    fig.update_layout(margin=dict(l=160, r=40, t=26, b=58), bargap=0.3)
    return fig


def _mud_correlation_colorscale() -> list:
    """Divergente azul↔rojo con gris neutro al centro: el punto medio debe leerse como
    «nada» y los polos como opuestos (nunca un arcoíris, nunca un tono en el centro)."""
    t = _mud_viz_tokens()
    return [[0.0, "#2a78d6"], [0.5, t["neutral"]], [1.0, "#e34948"]]


def format_num(val: float | int | None, digits: int = 2) -> str:
    if val is None or pd.isna(val):
        return "—"
    return f"{val:.{digits}f}"


def series_summary(series: pd.Series) -> str:
    return f"min {format_num(series.min())}, max {format_num(series.max())}, avg {format_num(series.mean())}"


def _render_chips_row(items: list[tuple[str, str]]) -> None:
    if not items:
        return
    try:
        cols = st.columns(len(items))
        for i, (label, color) in enumerate(items):
            with cols[i]:
                st.badge(label, color=color, width="content")
    except Exception:
        st.markdown(" ".join(f":{c}-badge[{l}]" for l, c in items))


def heatmap_numeric_stats(df: pd.DataFrame, cols: list) -> pd.DataFrame:
    cols = [c for c in cols if c in df.columns]
    rows = []
    for c in cols:
        s = _safe_numeric_series(df, c).dropna()
        if s.empty:
            rows.append({"Parámetro": str(c), "Mínimo": np.nan, "Promedio": np.nan, "Máximo": np.nan, "N": 0})
        else:
            rows.append({"Parámetro": str(c), "Mínimo": float(s.min()), "Promedio": float(s.mean()), "Máximo": float(s.max()), "N": int(len(s))})
    return pd.DataFrame(rows)


def stats_df_to_heatmap_chips(stats_df: pd.DataFrame, max_chips: int = 12) -> list[tuple[str, str]]:
    items = []
    for _, r in stats_df.iterrows():
        if int(r.get("N", 0) or 0) < 1:
            continue
        name = str(r["Parámetro"])
        if len(name) > 18:
            name = name[:16] + "…"
        lo, mid, hi = r["Mínimo"], r["Promedio"], r["Máximo"]
        sub = f"{format_num(lo, 1)}–{format_num(hi, 1)} · μ{format_num(mid, 1)}"
        items.append((f"{name}: {sub}", "blue"))
        if len(items) >= max_chips:
            break
    return items


def build_minmax_mean_spine_figure(stats_df: pd.DataFrame, title: str = "Rango por parámetro") -> go.Figure | None:
    if stats_df is None or stats_df.empty or "Parámetro" not in stats_df.columns:
        return None
    fig = go.Figure()
    for _, r in stats_df.iterrows():
        lo, mid, hi = r.get("Mínimo"), r.get("Promedio"), r.get("Máximo")
        p = str(r["Parámetro"])
        if pd.isna(lo) or pd.isna(mid) or pd.isna(hi):
            continue
        lo_f, mid_f, hi_f = float(lo), float(mid), float(hi)
        span = hi_f - lo_f
        ym = 0.5 if span <= 0 or not np.isfinite(span) else float(min(1.0, max(0.0, (mid_f - lo_f) / span)))
        fig.add_trace(go.Scatter(x=[p, p], y=[0.0, 1.0], mode="lines", line=dict(width=3, color="rgba(148,163,184,0.9)"), showlegend=False))
        fig.add_trace(go.Scatter(x=[p], y=[ym], mode="markers", marker=dict(size=11, color="#0ea5e9"), showlegend=False))
    fig.update_layout(title=dict(text=title, x=0.02, xanchor="left"), height=400, template=PLOTLY_TEMPLATE)
    return fig


def build_hist_with_trend(values, title: str, x_label: str, nbins: int = 30) -> go.Figure:
    vals = pd.Series(values).dropna()
    if vals.empty:
        return go.Figure()
    return px.histogram(vals, nbins=nbins, title=title, labels={"value": x_label})


# Equivalencia columna de la bitácora <-> etiqueta del reporte. Vive aquí para que el
# origen de cada número sea consultable desde la app y no haya que ir al código.
# Primera columna: encabezado en español y, entre paréntesis, el del formato inglés, para
# que la tabla sirva con cualquiera de los dos ajustes de idioma.
MUD_COLUMN_GLOSSARY: list[tuple[str, str, str]] = [
    ("Profundidad / TVD  (Depth MD / TVD)", "Profundidad / TVD", ""),
    ("Densidad @ N°C  (D @ N°C)", "Densidad / Temp.", "N es la temperatura a la que se midió esa muestra"),
    ("Visc. Embudo @ N°C  (Fv @ N°C)", "Visc. Embudo", "el reporte no da temperatura propia del embudo: se usa la de la densidad"),
    ("VP @ 65°C  (PV @ 65°C)", "VP", "medido a la Temp. Reología del reporte"),
    ("PC  (YP)", "PC", ""),
    ("Gel 10s / 10m / 30m  (GELS)", "10s/10m/30m Gel", ""),
    ("R600 … R3  (Lectura 600 … 3)", "R600 / R300, R200 / R100, R6 / R3", ""),
    ("Filtrado HTHP @ 149°C  (HTHP @ 149°C)", "Filtrado HTHP", "149 °C es la condición estándar del ensayo"),
    ("Revoque HTHP  (Cake (HTHP))", "Revoque API / HTHP", "se toma el valor de HTHP, no el de API"),
    ("Solidos Corregidos  (Corr Solid)", "Solidos Corregidos", ""),
    ("Aceite  (NAP)", "Aceite %vol", "NAP = fase no acuosa; en lodo base aceite es el aceite"),
    ("Agua  (Water)", "Agua %vol", ""),
    ("Aceite / Agua (Rel. A/A)  (NAP 2 / Water Ratio)", "Aceite / Agua", "las dos partes de la relación: 87 / 13 → 87 y 13"),
    ("Arena  (Sand)", "Arena", ""),
    ("Cloruros en Lodo  (Chlorides)", "Cloruros en Lodo", "mg/L"),
    ("Cloruros Fase Acuosa  (Water Phase Salinity)", "Cloruros Fase Acuosa", "en mg/L, tal como lo da el reporte"),
    ("Salinidad", "Salinidad", "%wt"),
    ("Alcalinidad (Pom)", "Alcalinidad (Pom)", ""),
    ("Exc. Cal  (Excess Lime)", "Exc. Cal", "kg/m³"),
    ("Estabilidad Elec.  (Elec. Stability)", "Estabilidad Elec.", "volt"),
    ("LGS TOTALES  (LGS)", "LGS TOTALES", "valor propio de cada muestra"),
    ("Solidos Alta Gravedad  (HGS)", "panel ANALISIS DE SOLIDOS → High Gravity %", ""),
    ("Peso Baja / Alta Gravedad  (LGS 2 · HGS 2)", "panel ANALISIS DE SOLIDOS → Low / High Gravity Wt.", "kg/m³"),
    ("SG Promedio Solidos  (ASG)", "panel ANALISIS DE SOLIDOS → Average SG Solids", ""),
    ("— los cuatro anteriores —", "panel ANALISIS DE SOLIDOS", "ese panel corresponde solo a la muestra de cierre del reporte, por eso las otras filas van vacías"),
    ("Calcio · CaCl2 · Tauy · n · k · Ensayo de Asentamiento (VSST)", "—", "el reporte Mi SWACO no los trae; quedan vacías"),
]


def _sanitize_filename(value: str, default: str = "mud_bitacora") -> str:
    value = (value or "").strip()
    if not value:
        return default
    value = re.sub(r"[^A-Za-z0-9_.-]+", "_", value)
    value = value.strip("._-")
    return value or default


def _default_mud_bitacora_basename(bitacora: pd.DataFrame) -> str:
    date_label = ""
    try:
        if "Date" in bitacora.columns and bitacora["Date"].notna().any():
            dmax = pd.to_datetime(bitacora["Date"], errors="coerce").dropna().max()
            if pd.notna(dmax):
                date_label = dmax.strftime("%Y-%m-%d")
    except Exception:
        pass
    return f"mud_bitacora_{date_label}" if date_label else "mud_bitacora"

# Mud Report – bitácora de propiedades de fluidos
# =========================
# Aliases por propiedad canónica (nombre en reporte -> clave bitácora)
MUD_PROPERTY_ALIASES = {
    "Density": ["density", "densidad", "mw", "mw (g/l)", "density @ c", "density @", "mud weight", "peso lodo", "densidad sp.gr"],
    "Marsh": ["marsh", "visc. marsh", "viscosidad marsh"],
    "Temperature": ["temperatura salida", "temp. de salida", "temperatura", "temp. de analisis", "temp. de análisis", "temp de analisis"],
    "VA": ["va", "visc.aparente", "visc. aparente", "viscosidad aparente"],
    "FV": ["fv", "fv @ c", "fv @ °c", "funnel viscosity", "viscosidad embudo"],
    "PV": ["pv", "pv (cp)", "pv @ c", "pv @ °c", "plastic viscosity", "viscosidad plástica", "viscoplastic", "visc. plastica", "visc.plastica"],
    "YP": ["yp", "yv", "yield point", "punto de cedencia", "yp (lb/100ft2)", "lb/100ft²", "pc"],
    "Gel_10s": ["gel 10s", "gel 10s/10m/30m", "gels 10s", "10s", "gel (10s)", "gel 10s/10m", "geles"],
    "Gel_10min": ["gel 10m", "gel 10min", "10min", "10m"],
    "Gel_30min": ["gel 30m", "gel 30min", "30min", "30m"],
    "L600": ["lectura 600", "l600"],
    "L300": ["lectura 300", "l300"],
    "L200": ["lectura 200", "l200"],
    "L100": ["lectura 100", "l100"],
    "L6": ["lectura 6", "l6"],
    "L3": ["lectura 3", "l3"],
    "Filtrado": ["filtrado", "filtrate", "fl temp", "hthp", "api filtrate", "fluid loss", "cake (hthp)", "filtrado hpht", "filtrado apat"],
    "Enjarre": ["enjarre", "cake"],
    "LGS": ["lgs", "lgs/hgs", "low gravity solids", "lgs (%)"],
    "HGS": ["hgs", "high gravity solids", "hgs (%)"],
    "Chlorides": ["chlorides", "cloruros", "chlorides (ppm)", "chlorides / calcium"],
    "Solids": ["solids", "corr solid", "solids content %", "sand %", "% sólidos", "% solidos", "sólidos no corregidos", "solidos no corregidos", "solidos corregidos"],
    "Oil": ["% aceite", "aceite %vol", "%oil", "oil"],
    "Water": ["% agua", "agua no correg", "%water", "water"],
    "RAA": ["raa", "r. aceite / agua", "rel. aceite/agua", "aceite/agua"],
    "AgNO3": ["agno3"],
    "Salinity": ["salinidad"],
    "Electrical_Stability": ["est. electrica", "estabilidad", "est. elect", "elec. stability"],
    "Alkalinity": ["alcalinidad"],
    "Excess_Cal": ["exceso de cal", "exc.cal", "exc cal"],
}
MUD_CANONICAL_ORDER = [
    "Date", "DateTime", "Properties", "Depth (MD)", "Depth (TVD)", "Fluid set", "Source", "Well", "Time", "FL Temp",
    "Density @ °C", "FV", "FV Temp", "FV @ °C", "PV", "PV Temp", "PV @ °C", "YP",
    "Gel_10s", "Gel_10min", "Gel_30min", "tau0",
    "L600", "L300", "L200", "L100", "L6", "L3",
    "HTHP", "HTHP @ °C", "Corr Solid", "NAP", "Water", "NAP Ratio", "Water Ratio",
    "Sand", "Cake (HTHP)", "Chlorides", "Calcium", "CaCl2", "Water Phase Salinity",
    "NaCL (Sol/Insol)", "Excess Lime", "Electrical_Stability",
    "LGS (%)", "HGS (%)", "LGS (kg/m³)", "HGS (kg/m³)", "ASG",
    "Additional Properties", "n (HB)", "K (HB)", "Viscometer Sag Shoe Test", "(VSST)",
    "Marsh", "Temperature", "VA", "Filtrado", "Enjarre", "LGS", "HGS", "Solids", "Oil", "RAA",
    "AgNO3", "Salinity", "Alkalinity", "Excess_Cal",
]
MUD_METADATA_COLUMNS = {
    "Date", "DateTime", "Properties", "Depth (MD)", "Depth (TVD)", "Fluid set", "Source", "Time",
    "Additional Properties", "Well",
}
MUD_ANALYTIC_EXCLUDE = {
    "DateTime", "Properties", "Depth (MD)", "Depth (TVD)", "Fluid set", "Source", "Time",
    "Density @ °C", "FV @ °C", "PV @ °C",
    # 'Additional Properties' es el índice de muestra del día, no una propiedad del lodo:
    # graficarlo o promediarlo no significa nada.
    "Additional Properties", "Well",
}
MUD_EXPORT_HEADER_SPECS = [
    ("Depth (MD)", "Depth (MD)", "m"),
    ("Depth (TVD)", "Depth (TVD)", "m"),
    ("Properties", "Properties", "N°"),
    ("Fluid set", "Fluid set", "Fluid"),
    ("Source", "Source", "Source"),
    ("Time", "Time", "time"),
    ("DateTime", "DateTime", "YYYY-MM-DDTHH:MM:SS"),
    ("FL Temp", "FL Temp", "°C"),
    ("D @ 54°C", "D @ 54°C", "kg/m³"),
    ("D @ 45°C", "D @ 45°C", "kg/m³"),
    ("D @ 44°C", "D @ 44°C", "kg/m³"),
    ("Fv @ 54°C", "Fv @ 54°C", "s/qt"),
    ("Fv @ 45°C", "Fv @ 45°C", "s/qt"),
    ("Fv @ 44°C", "Fv @ 44°C", "s/qt"),
    ("PV @ 65°C", "PV @ 65°C", "cP"),
    ("YP", "YP", "lb/100ft²"),
    ("Gel_10s", "GELS 10s", "lb/100ft²"),
    ("Gel_10min", "GELS 10min", "lb/100ft²"),
    ("Gel_30min", "GELS 30min", "lb/100ft²"),
    ("tau0", "tau0", "lb/100ft²"),
    ("L600", "Lectura 600", "lb/100 ft2"),
    ("L300", "Lectura 300", "lb/100 ft2"),
    ("L200", "Lectura 200", "lb/100 ft2"),
    ("L100", "Lectura 100", "lb/100 ft2"),
    ("L6", "Lectura 6", "lb/100 ft2"),
    ("L3", "Lectura 3", "lb/100 ft2"),
    ("HTHP @ 149°C", "HTHP @ 149°C", "HTHP"),
    ("Corr Solid", "Corr Solid", "%"),
    ("NAP", "NAP", "%"),
    ("Water", "Water", "%"),
    ("NAP 2", "NAP 2", "%"),
    ("Water Ratio", "Water Ratio", "%"),
    ("Sand", "Sand", "%"),
    ("Cake (HTHP)", "Cake (HTHP)", "32nd"),
    ("Chlorides", "Chlorides", "mg/L"),
    ("Calcium", "Calcium", "mg/L"),
    ("CaCl2", "CaCl2", "mg/L"),
    # El reporte da los cloruros de la fase acuosa en mg/L, no en ppm: en una salmuera
    # de ~1,2 g/cm³ no son equivalentes, así que la unidad dice lo que realmente se copió.
    ("Water Phase Salinity", "Water Phase Salinity", "mg/L"),
    ("Salinity", "Salinidad", "%wt"),
    ("Alkalinity", "Alcalinidad (Pom)", "cm³"),
    ("Excess Lime", "Excess Lime", "kg/m³"),
    ("Electrical_Stability", "Elec. Stability", "V"),
    ("LGS (%)", "LGS", "%"),
    ("HGS (%)", "HGS", "%"),
    ("LGS (kg/m³)", "LGS 2", "kg/m³"),
    ("HGS (kg/m³)", "HGS 2", "kg/m³"),
    ("ASG", "ASG", "SG"),
    ("Additional Properties", "Additional Properties", "Properties"),
    ("n (HB)", "n (HB)", "dec"),
    ("K (HB)", "K (HB)", "lb*s^n'/100ft2"),
    ("Viscometer Sag Shoe Test", "Viscometer Sag Shoe Test", "lbm/gal"),
]


MUD_LANG_ES = "es"
MUD_LANG_EN = "en"

# Encabezados en español: la etiqueta y la unidad TAL COMO las escribe el reporte de lodo,
# no una traducción libre. MUD_EXPORT_HEADER_SPECS viene del formato WellSight, que es un
# reporte en inglés; con un reporte en español esos nombres (NAP, Corr Solid, Cake, Excess
# Lime, Elec. Stability…) no aparecen en ninguna parte del original.
# Nota: las etiquetas del panel de sólidos y del de reología están en inglés en el propio
# reporte en español (High Gravity, Average SG Solids, Tauy, n, k), así que se dejan así.
MUD_HEADERS_ES: dict[str, tuple[str, str]] = {
    "Depth (MD)": ("Profundidad", "m"),
    "Depth (TVD)": ("TVD", "m"),
    "Properties": ("Propiedades", "N°"),
    "Fluid set": ("Tipo de Lodo", "Lodo"),
    "Source": ("Muestra", "Origen"),
    "Time": ("Hora", "hora"),
    "DateTime": ("Fecha y hora", "AAAA-MM-DDTHH:MM:SS"),
    "FL Temp": ("Temp. Flow Line", "°C"),
    "PV @ 65°C": ("VP @ 65°C", "cP"),
    "YP": ("PC", "lbf/100ft²"),
    "Gel_10s": ("Gel 10s", "lbf/100ft²"),
    "Gel_10min": ("Gel 10m", "lbf/100ft²"),
    "Gel_30min": ("Gel 30m", "lbf/100ft²"),
    "tau0": ("Tauy", "lbf/100ft²"),
    "L600": ("R600", "lbf/100ft²"),
    "L300": ("R300", "lbf/100ft²"),
    "L200": ("R200", "lbf/100ft²"),
    "L100": ("R100", "lbf/100ft²"),
    "L6": ("R6", "lbf/100ft²"),
    "L3": ("R3", "lbf/100ft²"),
    "HTHP @ 149°C": ("Filtrado HTHP @ 149°C", "cc/30min"),
    "Corr Solid": ("Solidos Corregidos", "%vol"),
    "NAP": ("Aceite", "%vol"),
    "Water": ("Agua", "%vol"),
    "NAP 2": ("Aceite (Rel. A/A)", "%"),
    "Water Ratio": ("Agua (Rel. A/A)", "%"),
    "Sand": ("Arena", "%vol"),
    "Cake (HTHP)": ("Revoque HTHP", '1/32"'),
    "Chlorides": ("Cloruros en Lodo", "mg/L"),
    "Calcium": ("Calcio", "mg/L"),
    "CaCl2": ("CaCl2", "mg/L"),
    "Water Phase Salinity": ("Cloruros Fase Acuosa", "mg/L"),
    "Salinity": ("Salinidad", "%wt"),
    "Alkalinity": ("Alcalinidad (Pom)", "cm³"),
    "Excess Lime": ("Exc. Cal", "kg/m³"),
    "Electrical_Stability": ("Estabilidad Elec.", "volt"),
    "LGS (%)": ("LGS TOTALES", "%"),
    # El panel ANALISIS DE SOLIDOS del reporte está rotulado en inglés (High Gravity,
    # Low Gravity Wt., Average SG Solids) incluso en la versión en español; aquí se
    # traducen porque la bitácora se lee en español. 'Wt.' es peso/concentración, y así
    # se distingue el porcentaje de la concentración en kg/m³.
    "HGS (%)": ("Solidos Alta Gravedad", "%"),
    "LGS (kg/m³)": ("Peso Baja Gravedad", "kg/m³"),
    "HGS (kg/m³)": ("Peso Alta Gravedad", "kg/m³"),
    "ASG": ("SG Promedio Solidos", "SG"),
    "Additional Properties": ("Propiedades Adicionales", "Propiedades"),
    "n (HB)": ("n", "dec"),
    "K (HB)": ("k", "lb*s^n'/100ft2"),
    "Viscometer Sag Shoe Test": ("Ensayo de Asentamiento (VSST)", "lbm/gal"),
}


def _mud_localize_specs(specs: list[tuple[str, str, str]], lang: str) -> list[tuple[str, str, str]]:
    """Traduce encabezado y unidad. Las claves de columna son internas y no cambian."""
    if lang != MUD_LANG_ES:
        return specs
    out: list[tuple[str, str, str]] = []
    for col, h1, h2 in specs:
        # Densidad y viscosidad de embudo llevan la temperatura medida en el nombre.
        m = re.fullmatch(r"(D|Fv) @ ([\d.]+)°C", col)
        if m:
            es_base, es_unit = ("Densidad", "kg/m³") if m.group(1) == "D" else ("Visc. Embudo", "sec/qt")
            out.append((col, f"{es_base} @ {m.group(2)}°C", es_unit))
            continue
        es = MUD_HEADERS_ES.get(col)
        out.append((col, es[0], es[1]) if es else (col, h1, h2))
    return out


def _mud_bitacora_title(lang: str, title_date: str) -> str:
    if lang == MUD_LANG_ES:
        base = "Reporte Diario de Propiedades del Fluido"
        return f"{base} — Reporte: {title_date}" if title_date else base
    base = "Daily Fluid Properties Daily Report"
    return f"{base} Report: {title_date}" if title_date else base


def _normalize_mud_property_name(label: str) -> str | None:
    """Mapea etiqueta de reporte a nombre canónico."""
    if not label or not isinstance(label, str):
        return None
    key = str(label).strip().lower()
    key = re.sub(r"\s+", " ", key)
    for canonical, aliases in MUD_PROPERTY_ALIASES.items():
        for a in aliases:
            # Límite de palabra, no substring suelto: alias cortos como "oil" o "pc"
            # no deben matchear dentro de texto no relacionado ("GASOIL", "PCN.Nq...").
            if re.search(rf"\b{re.escape(a)}\b", key) or re.search(rf"\b{re.escape(key)}\b", a):
                return canonical
    if "gel" in key and "10s" in key:
        return "Gel_10s"
    if "gel" in key and "10" in key and "30" not in key:
        return "Gel_10min"
    if "gel" in key and "30" in key:
        return "Gel_30min"
    return None


def _extract_numeric(val) -> float | None:
    """Extrae un número de una celda o string libre."""
    if val is None or (isinstance(val, float) and np.isnan(val)):
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if not s:
        return None
    s = s.replace(" ", " ")
    # quitar miles y normalizar decimales
    s = re.sub(r"(?<=\d),(?=\d{3}(?:\D|$))", "", s)
    s = s.replace(",", ".")
    m = re.search(r"[-+]?\d*\.?\d+", s)
    if m:
        try:
            return float(m.group(0))
        except ValueError:
            return None
    return None


def _extract_all_numbers(val) -> list[float]:
    if val is None or (isinstance(val, float) and np.isnan(val)):
        return []
    s = str(val).strip().replace(" ", " ")
    if not s:
        return []
    s = re.sub(r"(?<=\d),(?=\d{3}(?:\D|$))", "", s)
    s = s.replace(",", ".")
    nums = []
    for tok in re.findall(r"[-+]?\d*\.?\d+", s):
        try:
            nums.append(float(tok))
        except ValueError:
            pass
    return nums


def _parse_gel_triple(val) -> tuple[float | None, float | None, float | None]:
    """Parsea '10/15/17' -> (10, 15, 17) y variantes '8/16/22'."""
    nums = _extract_all_numbers(val)
    if len(nums) >= 3:
        return nums[0], nums[1], nums[2]
    if len(nums) == 2:
        return nums[0], nums[1], None
    if len(nums) == 1:
        return nums[0], None, None
    return None, None, None


def _extract_date_from_text(text: str) -> pd.Timestamp | None:
    if not text:
        return None
    month_map = {
        "ene": 1, "feb": 2, "mar": 3, "abr": 4, "may": 5, "jun": 6,
        "jul": 7, "ago": 8, "sep": 9, "oct": 10, "nov": 11, "dic": 12,
    }
    # yyyy-mm-dd / yyyy.mm.dd / yyyy/mm/dd
    m = re.search(r"((?:19|20)\d{2})[./-](\d{1,2})[./-](\d{1,2})", text)
    if m:
        y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
        try:
            return pd.Timestamp(y, mo, d).normalize()
        except ValueError:
            pass
    # dd-mm-yyyy / dd/mm/yyyy
    m = re.search(r"(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})", text)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if y < 100:
            y += 2000
        try:
            return pd.Timestamp(y, mo, d).normalize()
        except ValueError:
            pass
    m = re.search(r"(\d{1,2})-([A-Za-z]{3})-(\d{2,4})", text)
    if m:
        d = int(m.group(1))
        mo = month_map.get(m.group(2).lower()[:3])
        y = int(m.group(3))
        if y < 100:
            y += 2000
        if mo:
            try:
                return pd.Timestamp(y, mo, d).normalize()
            except ValueError:
                pass
    return None


def _date_from_filename_or_today(name: str) -> pd.Timestamp:
    """Extrae fecha de nombre de archivo o usa hoy."""
    if not name:
        return pd.Timestamp.now().normalize()
    d = _extract_date_from_text(name)
    return d if d is not None else pd.Timestamp.now().normalize()


def _mud_num_to_text(val) -> str:
    num = _extract_numeric(val)
    if num is None:
        s = str(val).strip() if val is not None else ""
        return s
    if abs(num - round(num)) < 1e-9:
        return str(int(round(num)))
    return f"{num:.12g}"


def _mud_clean_cell_text(val) -> str:
    if val is None or (isinstance(val, float) and np.isnan(val)):
        return ""
    if isinstance(val, pd.Timestamp):
        return val.strftime("%Y-%m-%d %H:%M:%S")
    try:
        import datetime as _dt
        if isinstance(val, _dt.time):
            return val.strftime("%H:%M")
    except Exception:
        pass
    return str(val).strip().replace("\u00a0", " ")


def _mud_parse_time_value(val):
    s = _mud_clean_cell_text(val)
    if not s:
        return None
    try:
        import datetime as _dt
        if hasattr(val, "hour") and hasattr(val, "minute") and not isinstance(val, pd.Timestamp):
            return _dt.time(val.hour, val.minute, getattr(val, "second", 0))
    except Exception:
        pass
    for fmt in ("%H:%M:%S", "%H:%M"):
        try:
            return datetime.strptime(s, fmt).time()
        except Exception:
            pass
    dt = pd.to_datetime(s, errors="coerce")
    if pd.notna(dt):
        return dt.time()
    return None


def _mud_compose_datetime(date_value, time_value) -> pd.Timestamp | None:
    base_date = pd.to_datetime(date_value, errors="coerce")
    if pd.isna(base_date):
        return None
    t = _mud_parse_time_value(time_value)
    if t is None:
        return base_date
    return pd.Timestamp.combine(base_date.normalize().date(), t)


def _mud_isoformat_no_tz(ts) -> str:
    ts = pd.to_datetime(ts, errors="coerce")
    if pd.isna(ts):
        return ""
    return ts.strftime("%Y-%m-%dT%H:%M:%S")


def _mud_pair_string(raw_value) -> str:
    nums = _extract_all_numbers(raw_value)
    if len(nums) >= 2:
        return f"{_mud_num_to_text(nums[0])} @ {_mud_num_to_text(nums[1])}"
    if len(nums) == 1:
        return _mud_num_to_text(nums[0])
    return _mud_clean_cell_text(raw_value)


def _mud_apply_daily_property(row_record: dict, label: str, unit: str, raw_value) -> None:
    low = re.sub(r"\s+", " ", _mud_clean_cell_text(label).lower())
    unit_low = re.sub(r"\s+", " ", _mud_clean_cell_text(unit).lower())
    nums = _extract_all_numbers(raw_value)
    if not low:
        return
    if low.startswith("depth"):
        if len(nums) >= 1:
            row_record["Depth (MD)"] = nums[0]
        if len(nums) >= 2:
            row_record["Depth (TVD)"] = nums[1]
        return
    if low.startswith("fl temp"):
        if nums:
            row_record["FL Temp"] = nums[0]
        return
    if low.startswith("d @") or (low.startswith("density") and "@" in low):
        temp_match = re.search(r"@\s*(\d+)", low)
        if temp_match and nums:
            row_record[f"D @ {temp_match.group(1)}°C"] = nums[0]
        row_record["Density @ °C"] = _mud_pair_string(raw_value)
        if nums:
            row_record["Density"] = nums[0]
        if len(nums) >= 2:
            row_record["Density Temp"] = nums[1]
        elif temp_match:
            row_record["Density Temp"] = float(temp_match.group(1))
        return
    if low.startswith("density"):
        row_record["Density @ °C"] = _mud_pair_string(raw_value)
        if nums:
            row_record["Density"] = nums[0]
        if len(nums) >= 2:
            row_record["Density Temp"] = nums[1]
        return
    if low.startswith("fv @"):
        temp_match = re.search(r"@\s*(\d+)", low)
        if temp_match and nums:
            row_record[f"Fv @ {temp_match.group(1)}°C"] = nums[0]
        if nums:
            row_record["FV"] = nums[0]
        if len(nums) >= 2:
            row_record["FV Temp"] = nums[1]
        elif temp_match:
            row_record["FV Temp"] = float(temp_match.group(1))
        row_record["FV @ °C"] = _mud_pair_string(raw_value)
        return
    if low.startswith("pv @"):
        if nums:
            row_record["PV"] = nums[0]
            row_record["PV @ 65°C"] = nums[0]
        if len(nums) >= 2:
            row_record["PV Temp"] = nums[1]
        row_record["PV @ °C"] = _mud_pair_string(raw_value)
        return
    if low == "yp" or low.startswith("yp "):
        if nums:
            row_record["YP"] = nums[0]
        return
    if low.startswith("gels"):
        g1, g2, g3 = _parse_gel_triple(raw_value)
        if g1 is not None:
            row_record["Gel_10s"] = g1
        if g2 is not None:
            row_record["Gel_10min"] = g2
        if g3 is not None:
            row_record["Gel_30min"] = g3
        return
    if low == "tau0":
        if nums:
            row_record["tau0"] = nums[0]
        return
    if low.startswith("600/300"):
        if len(nums) >= 1:
            row_record["L600"] = nums[0]
        if len(nums) >= 2:
            row_record["L300"] = nums[1]
        return
    if low.startswith("200/100"):
        if len(nums) >= 1:
            row_record["L200"] = nums[0]
        if len(nums) >= 2:
            row_record["L100"] = nums[1]
        return
    if low.startswith("6/3"):
        if len(nums) >= 1:
            row_record["L6"] = nums[0]
        if len(nums) >= 2:
            row_record["L3"] = nums[1]
        return
    if low.startswith("hthp"):
        if nums:
            row_record["HTHP"] = nums[0]
            row_record["HTHP @ 149°C"] = nums[0]
        if len(nums) >= 2:
            row_record["HTHP @ °C"] = nums[1]
        return
    if low.startswith("corr solid"):
        if nums:
            row_record["Corr Solid"] = nums[0]
        return
    if low.startswith("nap / water ratio"):
        if len(nums) >= 1:
            row_record["NAP Ratio"] = nums[0]
            row_record["NAP 2"] = nums[0]
        if len(nums) >= 2:
            row_record["Water Ratio"] = nums[1]
        return
    if low.startswith("nap / water"):
        if len(nums) >= 1:
            row_record["NAP"] = nums[0]
        if len(nums) >= 2:
            row_record["Water"] = nums[1]
        return
    if low == "sand" or low.startswith("sand "):
        if nums:
            row_record["Sand"] = nums[0]
        return
    if low.startswith("cake"):
        if nums:
            row_record["Cake (HTHP)"] = nums[0]
        return
    if low.startswith("chlorides / calcium"):
        if len(nums) >= 1:
            row_record["Chlorides"] = nums[0]
        if len(nums) >= 2:
            row_record["Calcium"] = nums[1]
        return
    if low == "cacl2" or low.startswith("cacl2 "):
        if nums:
            row_record["CaCl2"] = nums[0]
        return
    if low.startswith("water phase salinity"):
        if nums:
            row_record["Water Phase Salinity"] = nums[0]
        return
    if low.startswith("nacl"):
        txt = _mud_clean_cell_text(raw_value)
        if txt and txt != "/":
            row_record["NaCL (Sol/Insol)"] = txt
        return
    if low.startswith("excess lime"):
        if nums:
            row_record["Excess Lime"] = nums[0]
        return
    if low.startswith("elec. stability"):
        if nums:
            row_record["Electrical_Stability"] = nums[0]
        return
    if low.startswith("lgs / hgs"):
        if len(nums) >= 1:
            target_lgs = "LGS (kg/m³)" if "kg/" in unit_low else "LGS (%)"
            row_record[target_lgs] = nums[0]
        if len(nums) >= 2:
            target_hgs = "HGS (kg/m³)" if "kg/" in unit_low else "HGS (%)"
            row_record[target_hgs] = nums[1]
        return
    if low == "asg":
        if nums:
            row_record["ASG"] = nums[0]
        return
    if low.startswith("n (hb)"):
        if nums:
            row_record["n (HB)"] = nums[0]
        return
    if low.startswith("k (hb)"):
        if nums:
            row_record["K (HB)"] = nums[0]
        return
    if low.startswith("viscometer sag shoe test"):
        if nums:
            row_record["Viscometer Sag Shoe Test"] = nums[0]
        return
    if low.startswith("(vsst)"):
        if nums:
            row_record["(VSST)"] = nums[0]
        else:
            txt = _mud_clean_cell_text(raw_value)
            if txt:
                row_record["(VSST)"] = txt
        return
    canonical = _normalize_mud_property_name(label)
    if canonical:
        _mud_apply_canonical_value(row_record, canonical, raw_value)


def _parse_mud_daily_report_sheet(df_raw: pd.DataFrame, source_name: str = "") -> list[dict]:
    df = df_raw.copy()
    if df.empty or df.shape[0] < 6 or df.shape[1] < 5:
        return []
    cell_a1 = _mud_clean_cell_text(df.iat[0, 0]) if df.shape[0] > 0 else ""
    cell_a2 = _mud_clean_cell_text(df.iat[1, 0]) if df.shape[0] > 1 else ""
    cell_a5 = _mud_clean_cell_text(df.iat[4, 0]) if df.shape[0] > 4 else ""
    if "daily fluid properties" not in cell_a1.lower() or "properties" not in cell_a2.lower() or "time" not in cell_a5.lower():
        return []

    report_date = _extract_date_from_text(cell_a1) or _date_from_filename_or_today(source_name)
    sample_cols = []
    for j in range(2, df.shape[1]):
        prop_txt = _mud_clean_cell_text(df.iat[1, j]) if df.shape[0] > 1 else ""
        fluid_txt = _mud_clean_cell_text(df.iat[2, j]) if df.shape[0] > 2 else ""
        src_txt = _mud_clean_cell_text(df.iat[3, j]) if df.shape[0] > 3 else ""
        time_txt = _mud_clean_cell_text(df.iat[4, j]) if df.shape[0] > 4 else ""
        if time_txt or fluid_txt or src_txt or (prop_txt and (fluid_txt or src_txt)):
            sample_cols.append(j)
    if not sample_cols:
        return []

    records: list[dict] = []
    for idx, j in enumerate(sample_cols, start=1):
        prop_id = _extract_numeric(df.iat[1, j]) if df.shape[0] > 1 else None
        time_raw = df.iat[4, j] if df.shape[0] > 4 else None
        ts = _mud_compose_datetime(report_date, time_raw)
        rec = {
            "Date": ts if ts is not None else report_date,
            "DateTime": _mud_isoformat_no_tz(ts if ts is not None else report_date),
            "Properties": int(prop_id) if prop_id is not None else idx,
            "Fluid set": _mud_clean_cell_text(df.iat[2, j]) if df.shape[0] > 2 else "",
            "Source": _mud_clean_cell_text(df.iat[3, j]) if df.shape[0] > 3 else source_name,
            "Time": _mud_parse_time_value(time_raw).strftime("%H:%M") if _mud_parse_time_value(time_raw) else _mud_clean_cell_text(time_raw),
            "Additional Properties": int(prop_id) if prop_id is not None else idx,
        }
        records.append(rec)

    row_texts = []
    for i in range(df.shape[0]):
        row_texts.append(" ".join(_mud_clean_cell_text(df.iat[i, c]) for c in range(df.shape[1]) if _mud_clean_cell_text(df.iat[i, c])))

    for i in range(5, df.shape[0]):
        label = _mud_clean_cell_text(df.iat[i, 0])
        if not label:
            continue
        unit = _mud_clean_cell_text(df.iat[i, 1]) if df.shape[1] > 1 else ""
        raw_vals = [df.iat[i, j] for j in sample_cols]
        whole_nums = _extract_all_numbers(row_texts[i])
        use_sequence = len(whole_nums) == len(sample_cols) and any(
            (not _mud_clean_cell_text(v)) or len(_extract_all_numbers(v)) != 1 for v in raw_vals
        )
        for idx, rec in enumerate(records):
            raw_value = whole_nums[idx] if use_sequence else raw_vals[idx]
            if not _mud_clean_cell_text(raw_value) and not isinstance(raw_value, (int, float)):
                continue
            _mud_apply_daily_property(rec, label, unit, raw_value)

    return [r for r in records if any(
        k not in MUD_METADATA_COLUMNS and pd.notna(v) and _mud_clean_cell_text(v) not in ("", "/")
        for k, v in r.items()
    )]


def _mud_apply_canonical_value(row_record: dict, canonical: str, raw_value) -> None:
    if canonical.startswith("Gel"):
        g1, g2, g3 = _parse_gel_triple(raw_value)
        if g1 is not None:
            row_record["Gel_10s"] = g1
        if g2 is not None:
            row_record["Gel_10min"] = g2
        if g3 is not None:
            row_record["Gel_30min"] = g3
        return
    if canonical == "RAA":
        nums = _extract_all_numbers(raw_value)
        if len(nums) >= 1:
            row_record[canonical] = nums[0]
        return
    num = _extract_numeric(raw_value)
    if num is not None:
        row_record[canonical] = num


def _parse_mud_text_block(text: str, row_record: dict) -> None:
    if not text:
        return
    text = text.replace("\u00a0", " ")
    if pd.isna(row_record.get("Date")) or row_record.get("Date") is None:
        d = _extract_date_from_text(text)
        if d is not None:
            row_record["Date"] = d

    patterns = [
        ("Density", [r"densidad[^\n\r:]*[: ]+([0-9.,]+)", r"density[^\n\r:]*[: ]+([0-9.,]+)"]),
        ("Marsh", [r"visc\.? marsh[^\n\r:]*[: ]+([0-9.,]+)", r"viscosidad marsh[^\n\r:]*[: ]+([0-9.,]+)"]),
        ("Temperature", [r"temperatura salida[^\n\r:]*[: ]+([0-9.,]+)", r"temp\. de salida[^\n\r:]*[: ]+([0-9.,]+)", r"temp\. de an[aá]lisis[^\n\r:]*[: ]+([0-9.,]+)"]),
        ("VA", [r"visc\.?aparente(?:\(va\))?[^\n\r:]*[: ]+([0-9.,]+)", r"visc\.? aparente[^\n\r:]*[: ]+([0-9.,]+)"]),
        ("PV", [r"visc\.?plastica(?:\(vp\))?[^\n\r:]*[: ]+([0-9.,]+)", r"visc\.? plastica[^\n\r:]*[: ]+([0-9.,]+)", r"\bPV\b[^\n\r:]*[: ]+([0-9.,]+)"]),
        ("YP", [r"punto cedente(?:\(yp\))?[^\n\r:]*[: ]+([0-9.,]+)", r"\bPC\b[^\n\r:]*[: ]+([0-9.,]+)", r"\bYP\b[^\n\r:]*[: ]+([0-9.,]+)"]),
        ("L600", [r"lectura 600[^\n\r:]*[: ]+([0-9.,]+)", r"l600[^\n\r:]*[: ]+([0-9.,]+)"]),
        ("L300", [r"lectura 300[^\n\r:]*[: ]+([0-9.,]+)", r"l300[^\n\r:]*[: ]+([0-9.,]+)"]),
        ("L200", [r"lectura 200[^\n\r:]*[: ]+([0-9.,]+)", r"l200[^\n\r:]*[: ]+([0-9.,]+)"]),
        ("L100", [r"lectura 100[^\n\r:]*[: ]+([0-9.,]+)", r"l100[^\n\r:]*[: ]+([0-9.,]+)"]),
        ("L6", [r"lectura 6[^\n\r:]*[: ]+([0-9.,]+)", r"l6[^\n\r:]*[: ]+([0-9.,]+)"]),
        ("L3", [r"lectura 3[^\n\r:]*[: ]+([0-9.,]+)", r"l3[^\n\r:]*[: ]+([0-9.,]+)"]),
        ("Filtrado", [r"filtrado hpht[^\n\r:]*[: ]+([0-9.,]+)", r"filtrado apat[^\n\r:]*[: ]+([0-9.,]+)", r"filtrado[^\n\r:]*[: ]+([0-9.,]+)"]),
        ("Enjarre", [r"enjarre[^\n\r:]*[: ]+([0-9.,]+)"]),
        ("Solids", [r"%\s*s[óo]lidos[^\n\r:]*[: ]+([0-9.,]+)", r"s[óo]lidos no corregidos[^\n\r:]*[: ]+([0-9.,]+)"]),
        ("Oil", [r"%\s*aceite[^\n\r:]*[: ]+([0-9.,]+)", r"aceite\s*%vol[^\n\r:]*[: ]+([0-9.,]+)"]),
        ("Water", [r"%\s*agua[^\n\r:]*[: ]+([0-9.,]+)", r"agua no correg[^\n\r:]*[: ]+([0-9.,]+)"]),
        ("RAA", [r"raa[^\n\r:]*[: ]+([0-9.,]+)", r"rel\. aceite/agua[^\n\r:]*[: ]+([0-9.,]+)", r"aceite/agua[^\n\r:]*[: ]+([0-9.,]+)"]),
        ("AgNO3", [r"agno3[^\n\r:]*[: ]+([0-9.,]+)"]),
        ("Chlorides", [r"cloruros[^\n\r:]*[: ]+([0-9.,]+)", r"chlorides[^\n\r:]*[: ]+([0-9.,]+)"]),
        ("Salinity", [r"salinidad[^\n\r:]*[: ]+([0-9.,]+)"]),
        ("Electrical_Stability", [r"est\.? electrica[^\n\r:]*[: ]+([0-9.,]+)", r"estabilidad[^\n\r:]*[: ]+([0-9.,]+)", r"est\.? elect\.?[^\n\r:]*[: ]+([0-9.,]+)"]),
        ("Alkalinity", [r"alcalinidad[^\n\r:]*[: ]+([0-9.,]+)"]),
        ("Excess_Cal", [r"exceso de cal[^\n\r:]*[: ]+([0-9.,]+)", r"exc\.cal[^\n\r:]*[: ]+([0-9.,]+)"]),
    ]
    for canonical, regexes in patterns:
        if canonical in row_record and pd.notna(row_record.get(canonical)):
            continue
        for pat in regexes:
            m = re.search(pat, text, flags=re.IGNORECASE)
            if m:
                _mud_apply_canonical_value(row_record, canonical, m.group(1))
                break

    m = re.search(r"geles?[^\n\r:]*[: ]+([0-9.,/ ]+)", text, flags=re.IGNORECASE)
    if m:
        _mud_apply_canonical_value(row_record, "Gel_10s", m.group(1))
    m = re.search(r"gel\s*10s/10m[^\n\r:]*[: ]+([0-9.,/ ]+)", text, flags=re.IGNORECASE)
    if m:
        _mud_apply_canonical_value(row_record, "Gel_10s", m.group(1))




def _parse_mud_lines(text: str, row_record: dict) -> None:
    if not text:
        return
    for raw_line in text.splitlines():
        line = (raw_line or "").strip()
        if not line:
            continue
        low = re.sub(r"\s+", " ", line.lower())
        nums = _extract_all_numbers(line)
        if not nums:
            continue

        if low.startswith("densidad") or low.startswith("density"):
            row_record["Density @ °C"] = _mud_pair_string(line)
            row_record["Density"] = nums[0] if nums else row_record.get("Density")
            if len(nums) >= 2:
                row_record["Density Temp"] = nums[1]
        elif low.startswith("visc. marsh") or low.startswith("viscosidad marsh"):
            row_record["Marsh"] = nums[-1]
        elif low.startswith("temperatura salida") or low.startswith("temp. de salida"):
            row_record["Temperature"] = nums[-1]
        elif low.startswith("temp. de analisis") or low.startswith("temp. de análisis"):
            row_record["Temperature"] = nums[-1]
        elif low.startswith("visc.aparente") or low.startswith("visc. aparente") or low.startswith("viscosidad aparente"):
            row_record["VA"] = nums[-1]
        elif low.startswith("visc.plastica") or low.startswith("visc. plastica") or low.startswith("pv ") or low == "pv":
            row_record["PV"] = nums[-1]
        elif low.startswith("punto cedente") or low.startswith("pc ") or low == "pc" or low.startswith("yp ") or low == "yp":
            row_record["YP"] = nums[-1]
        elif low.startswith("lectura 600"):
            row_record["L600"] = nums[0]
        elif low.startswith("lectura 300"):
            row_record["L300"] = nums[0]
        elif low.startswith("lectura 200"):
            row_record["L200"] = nums[0]
        elif low.startswith("lectura 100"):
            row_record["L100"] = nums[0]
        elif low.startswith("lectura 6"):
            row_record["L6"] = nums[0]
        elif low.startswith("lectura 3"):
            row_record["L3"] = nums[0]
        elif low.startswith("l600/l300") and len(nums) >= 2:
            row_record["L600"] = nums[0]
            row_record["L300"] = nums[1]
        elif low.startswith("l200/l100") and len(nums) >= 2:
            row_record["L200"] = nums[0]
            row_record["L100"] = nums[1]
        elif low.startswith("l6/l3") and len(nums) >= 2:
            row_record["L6"] = nums[0]
            row_record["L3"] = nums[1]
        elif low.startswith("filtrado hpht") or low.startswith("filtrado apat") or low.startswith("filtrado"):
            row_record["Filtrado"] = nums[-1]
        elif low.startswith("enjarre"):
            row_record["Enjarre"] = nums[-1]
        elif low.startswith("geles") or low.startswith("gel 10s/10m"):
            if len(nums) >= 1:
                row_record["Gel_10s"] = nums[0]
            if len(nums) >= 2:
                row_record["Gel_10min"] = nums[1]
            if len(nums) >= 3:
                row_record["Gel_30min"] = nums[2]
        elif low.startswith("% sólidos") or low.startswith("% solidos") or low.startswith("sólidos no corregidos") or low.startswith("solidos no corregidos"):
            row_record["Solids"] = nums[-1]
        elif low.startswith("% aceite") or low.startswith("aceite %vol"):
            row_record["Oil"] = nums[-1]
        elif low.startswith("% agua") or low.startswith("agua no correg"):
            row_record["Water"] = nums[-1]
        elif low.startswith("raa") or low.startswith("rel. aceite/agua") or low.startswith("aceite/agua"):
            row_record["RAA"] = nums[-2] if len(nums) >= 2 else nums[0]
        elif low.startswith("agno3"):
            row_record["AgNO3"] = nums[-1]
        elif low.startswith("cloruros"):
            row_record["Chlorides"] = nums[-1]
        elif low.startswith("salinidad"):
            row_record["Salinity"] = nums[-1]
        elif low.startswith("est. electrica") or low.startswith("estabilidad") or low.startswith("est. elect"):
            row_record["Electrical_Stability"] = nums[-1]
        elif low.startswith("alcalinidad"):
            row_record["Alkalinity"] = nums[-1]
        elif low.startswith("exceso de cal") or low.startswith("exc.cal"):
            row_record["Excess_Cal"] = nums[-1]
def _parse_mud_excel_sheet(df_raw: pd.DataFrame, source_name: str = "") -> list[dict]:
    """Parsea una hoja Excel de propiedades de lodo (formato filas propiedad/valor o tabla)."""
    daily_rows = _parse_mud_daily_report_sheet(df_raw, source_name)
    if daily_rows:
        return daily_rows

    out: list[dict] = []
    date = _date_from_filename_or_today(source_name)
    row_record: dict = {"Date": date, "Source": source_name}

    df = df_raw.copy()
    df.columns = [str(c).strip() for c in df.columns]
    prop_col = None
    for c in df.columns:
        cl = c.lower()
        if "propert" in cl or "parámetro" in cl or "parameter" in cl or c == "Unnamed: 0" or cl == "0":
            prop_col = c
            break
    if prop_col is None and len(df.columns) >= 1:
        prop_col = df.columns[0]

    if prop_col is not None:
        for _, r in df.iterrows():
            label = r.get(prop_col)
            if pd.isna(label):
                continue
            canonical = _normalize_mud_property_name(str(label))
            if canonical:
                for c in df.columns:
                    if c == prop_col:
                        continue
                    v = r.get(c)
                    if pd.isna(v):
                        continue
                    _mud_apply_canonical_value(row_record, canonical, v)
                    if canonical in row_record or canonical.startswith("Gel"):
                        break

    if not any(k for k in row_record if k not in ("Date", "Source")):
        for col in df.columns:
            canonical = _normalize_mud_property_name(col)
            if canonical:
                vals = df[col].dropna().tolist()
                if vals:
                    _mud_apply_canonical_value(row_record, canonical, vals[0])

    text_blob = "\n".join(
        " ".join(str(v) for v in row.tolist() if pd.notna(v))
        for _, row in df.iterrows()
    )
    _parse_mud_text_block(text_blob, row_record)
    _parse_mud_lines(text_blob, row_record)

    if any(k for k in row_record if k not in ("Date", "Source")):
        out.append(row_record)
    return out


def _parse_mud_csv(df_raw: pd.DataFrame, source_name: str = "") -> list[dict]:
    """Parsea CSV de propiedades de lodo (igual lógica que Excel)."""
    return _parse_mud_excel_sheet(df_raw, source_name)


def _pdf_index_row_centers(page) -> tuple[list[float], float | None]:
    """
    Ubica la primera fila de índices 1..N bajo 'Propiedades'/'Properties' (o
    'Propiedades Adicionales') y devuelve (centros x de cada columna, top de esa fila).
    El top permite descartar el texto de título por encima de la tabla (nombre de
    pozo/operador/software), que de otro modo se leería como una fila de propiedad.
    """
    words = page.extract_words()
    rows: dict = {}
    for w in words:
        rows.setdefault(round(w["top"], 1), []).append(w)
    for top in sorted(rows):
        ws = sorted(rows[top], key=lambda w: w["x0"])
        # Exige que la fila empiece con "Properties"/"Propiedades": reportes combinados
        # (p.ej. "Daily Drilling Fluid Report") traen otras listas numeradas antes en la
        # página (Pumps 1 2 3 4 5, etc.) que también matchean 1..N si no se filtra esto.
        if not ws or not re.match(r"^(propiedades|properties)$", ws[0]["text"], re.IGNORECASE):
            continue
        digit_words = [w for w in ws if re.fullmatch(r"\d", w["text"])]
        if len(digit_words) >= 2:
            texts = [w["text"] for w in digit_words]
            if texts == [str(i + 1) for i in range(len(texts))]:
                return [(w["x0"] + w["x1"]) / 2 for w in digit_words], top
    return [], None


def _pdf_group_chars_into_rows(page, y_tolerance: float = 3.0) -> list[list[dict]]:
    """Agrupa los caracteres del PDF en líneas según su posición vertical."""
    chars = sorted(page.chars, key=lambda c: c["top"])
    rows: list[list[dict]] = []
    for c in chars:
        if rows and abs(c["top"] - rows[-1][0]["top"]) <= y_tolerance:
            rows[-1].append(c)
        else:
            rows.append([c])
    for r in rows:
        r.sort(key=lambda c: c["x0"])
    return rows


def _split_row_chars_into_runs(chars: list[dict], big_gap: float = 10.0) -> list[list[dict]]:
    """
    Agrupa caracteres consecutivos (ordenados por x0) en 'runs' que representan un
    único valor de columna. Corta en un espacio real entre columnas (hueco > big_gap)
    y también cuando WellSight pega dos valores sin separador: ahí el siguiente número
    arranca con un x0 ligeramente MENOR al x1 del carácter anterior (solape de ~2px por
    cómo esa plantilla ajusta números demasiado anchos para la celda).
    """
    chars_sorted = sorted(chars, key=lambda c: c["x0"])
    runs: list[list[dict]] = []
    max_x1 = None
    for c in chars_sorted:
        gap = None if max_x1 is None else c["x0"] - max_x1
        if gap is None or gap < -0.5 or gap > big_gap:
            runs.append([c])
        else:
            runs[-1].append(c)
        max_x1 = c["x1"] if max_x1 is None else max(max_x1, c["x1"])

    # El mismo solape que separa columnas pegadas a veces dispara dos cortes seguidos
    # (deja un fragmento suelto de 1-2 caracteres pegado, sin espacio real, al run
    # siguiente). Ese fragmento es el dígito inicial del número que sigue y se
    # re-adjunta a él. No aplica si el hueco hacia el run siguiente es un espacio de
    # columna real (varias decenas de px): ahí un valor corto de 2 caracteres (p.ej.
    # "18") es un valor de columna legítimo, no un fragmento espurio.
    merged: list[list[dict]] = []
    i = 0
    while i < len(runs):
        run = runs[i]
        stripped_len = len("".join(c["text"] for c in run).strip())
        gap_to_next = runs[i + 1][0]["x0"] - max(c["x1"] for c in run) if i + 1 < len(runs) else None
        if 0 < stripped_len <= 2 and gap_to_next is not None and gap_to_next < 3:
            runs[i + 1] = run + runs[i + 1]
            i += 1
            continue
        merged.append(run)
        i += 1
    return merged


def _split_wide_run_by_nearest_center(run: list[dict], centers: list[float], width: float) -> list[list[dict]]:
    """
    Un run más ancho que ~1.4 columnas suele ser texto de varias columnas pegado sin
    ningún espacio entre ellas (p.ej. "BaraXcel-1 AislaBaraXcel-1 Aisla..." en columnas
    angostas) — a diferencia del solape de números en columnas anchas, aquí el hueco
    entre valores es exactamente 0, así que _split_row_chars_into_runs no lo separa.
    Se subdivide carácter por carácter por cercanía al centro de columna más próximo.
    """
    x0 = min(c["x0"] for c in run)
    x1 = max(c["x1"] for c in run)
    if (x1 - x0) <= width * 1.4:
        return [run]
    groups: dict[int, list[dict]] = {}
    for c in run:
        cx = (c["x0"] + c["x1"]) / 2
        idx = min(range(len(centers)), key=lambda i: abs(centers[i] - cx))
        groups.setdefault(idx, []).append(c)
    return [groups[k] for k in sorted(groups)]


def _bucket_chars_by_column(chars: list[dict], centers: list[float]) -> list[str]:
    """
    Reconstruye el texto de cada columna de muestra a partir de caracteres del PDF.
    Primero agrupa los caracteres en 'runs' contiguos (un run = un valor de columna,
    con sus espacios internos conservados, p.ej. "85.71 / 14.29"), y luego asigna cada
    run a la columna cuyo centro esté más cerca. Asignar por centroide de todo el run
    (en vez de carácter por carácter) tolera el pequeño desalineamiento que WellSight
    introduce cuando un número ocupa casi todo el ancho de la celda.
    Un run cuyo centro cae lejos de TODOS los centros (más de un ancho de columna) se
    descarta: en layouts con columnas angostas, la misma línea de texto puede traer un
    panel lateral (rango objetivo, comentario de actividad) que no es una muestra real.
    """
    if not centers:
        return []
    width = (centers[1] - centers[0]) if len(centers) > 1 else 80.0
    values = ["" for _ in centers]
    raw_runs = _split_row_chars_into_runs(chars)
    sub_runs = [sr for run in raw_runs for sr in _split_wide_run_by_nearest_center(run, centers, width)]
    for run in sub_runs:
        text = "".join(c["text"] for c in run).strip()
        if not text:
            continue
        run_center = (min(c["x0"] for c in run) + max(c["x1"] for c in run)) / 2
        idx = min(range(len(centers)), key=lambda i: abs(centers[i] - run_center))
        if abs(centers[idx] - run_center) > width:
            continue
        values[idx] = f"{values[idx]} {text}".strip() if values[idx] else text
    return values


def _pdf_extract_label_and_unit(row_chars: list[dict], unit_boundary: float, big_gap: float = 6.0) -> tuple[str, str]:
    """
    Separa etiqueta y unidad de la parte de la fila anterior a la columna de muestras,
    por hueco real entre palabras: el hueco etiqueta→unidad es grande (decenas de px)
    frente al hueco entre palabras de una misma etiqueta (~1-2px). No depende de una
    columna de unidad a ancho fijo, porque ese ancho cambia entre plantillas de PDF.
    """
    pre_value_chars = [c for c in row_chars if c["x0"] < unit_boundary]
    pre_runs = _split_row_chars_into_runs(pre_value_chars, big_gap=big_gap)
    if not pre_runs:
        return "", ""
    if len(pre_runs) == 1:
        return "".join(c["text"] for c in pre_runs[0]).strip(), ""
    label = " ".join("".join(c["text"] for c in r).strip() for r in pre_runs[:-1]).strip()
    unit = "".join(c["text"] for c in pre_runs[-1]).strip()
    return label, unit


def _pdf_detect_n_active_from_time_row(page, col_centers: list[float], page_header_top: float, unit_boundary: float) -> int:
    """
    Cuenta cuántas columnas de muestra tienen datos reales, usando la fila Time/Tiempo
    como referencia: sus valores ('04:00') son cortos y siempre quedan bien separados
    por un hueco real, a diferencia de campos de texto largos (Fluid Set, Source) que
    en columnas angostas pueden pegarse sin espacio entre una muestra y la siguiente y
    dar un conteo de columnas poco confiable.
    """
    for row_chars in _pdf_group_chars_into_rows(page):
        if row_chars[0]["top"] < page_header_top:
            continue
        label_part, _ = _pdf_extract_label_and_unit(row_chars, unit_boundary)
        if label_part.lower() in ("tiempo", "time"):
            value_chars = [c for c in row_chars if c["x0"] >= unit_boundary]
            values = _bucket_chars_by_column(value_chars, col_centers)
            n_active = sum(1 for v in values if v)
            if n_active:
                return n_active
    return 0


def _parse_mud_pdf_daily_report_grid(pdf, source_name: str = "") -> list[dict]:
    """
    Parser posicional para el layout WellSight 'Fluidos de Perforación - Reporte
    Diario de Propiedades del Fluido' / 'Daily Fluid Properties Daily Report' /
    'Daily Drilling Fluid Report'. Reconstruye la grilla (etiqueta, unidad, valor por
    muestra) a partir de las coordenadas del texto del PDF y reutiliza
    _mud_apply_daily_property para que la bitácora salga idéntica a la que genera
    _parse_mud_daily_report_sheet con Excel.
    """
    full_text = "\n".join(page.extract_text() or "" for page in pdf.pages)
    if not re.search(
        r"reporte\s+diario\s+de\s+propiedades\s+del\s+fluido|daily\s+fluid\s+properties|daily\s+drilling\s+fluid\s+report",
        full_text,
        re.IGNORECASE,
    ):
        return []

    report_date = _extract_date_from_text(full_text) or _date_from_filename_or_today(source_name)
    records: list[dict] | None = None
    label_buffer = ""
    # Hueco vertical entre filas dentro de la tabla de propiedades (~8-16px en los
    # layouts vistos). Un salto mucho mayor señala el fin de la tabla: reportes
    # combinados siguen con otra sección en la misma página (p.ej. "Fluid Volume
    # Breakdown" con nombres de producto que pueden matchear un alias de propiedad
    # por casualidad, como "GASOIL" conteniendo "oil").
    MAX_ROW_GAP = 40.0

    for page in pdf.pages:
        # No se hereda la tabla de una página a la siguiente: reportes combinados
        # traen después páginas de otro tipo ("Daily Concentration Report", "Daily
        # Inventory & Tickets") que no tienen encabezado "Properties" propio y
        # contaminarían la bitácora si se reusaran las posiciones de columna previas.
        col_centers, header_top = _pdf_index_row_centers(page)
        if not col_centers:
            continue
        page_header_top = header_top

        unit_boundary = col_centers[0] - (col_centers[1] - col_centers[0]) / 2 if len(col_centers) > 1 else col_centers[0] - 40
        n_active_hint = _pdf_detect_n_active_from_time_row(page, col_centers, page_header_top, unit_boundary)

        last_row_top: float | None = None
        for row_chars in _pdf_group_chars_into_rows(page):
            top = row_chars[0]["top"]
            # Descarta texto por encima de la tabla (título con nombre de pozo/operador/software):
            # de otro modo se leería como una fila de propiedad y contaminaría la bitácora.
            if top < page_header_top:
                continue

            label_part, unit = _pdf_extract_label_and_unit(row_chars, unit_boundary)
            value_chars = [c for c in row_chars if c["x0"] >= unit_boundary]

            if not label_part and not label_buffer:
                continue

            # Una fila de índices ("Propiedades Adicionales 1 2 3...") puede aparecer
            # tras un hueco grande sin que eso signifique fin de tabla: sigue siendo la
            # misma tabla de propiedades, solo con una sub-sección más abajo en la
            # página. El salto se exceptúa únicamente para esa fila puente.
            is_header_row = bool(re.match(r"^(propiedades|properties)\b", label_part.lower()))
            if last_row_top is not None and top - last_row_top > MAX_ROW_GAP and not is_header_row:
                break
            last_row_top = top

            if re.match(r"^(creado|created)\b", label_part, re.IGNORECASE):
                label_buffer = ""
                continue

            label = f"{label_buffer} {label_part}".strip() if label_buffer else label_part
            low_label = label.lower()

            # Fila de índices (Propiedades / Propiedades Adicionales): ya usada para centrar columnas.
            if is_header_row:
                label_buffer = ""
                continue

            has_content = bool(unit) or any(c["text"].strip() for c in value_chars)
            if not has_content:
                # Un fragmento entre paréntesis (p.ej. "(VSST)") es un sufijo final de la
                # etiqueta de ARRIBA, no el inicio de la etiqueta de la fila siguiente:
                # se descarta en vez de encolarlo como prefijo.
                label_buffer = "" if label_part.startswith("(") else label
                continue
            label_buffer = ""

            if records is None:
                n_now = n_active_hint
                if not n_now and low_label in ("set de fluido", "fluid set", "tiempo", "time"):
                    probe = _bucket_chars_by_column(value_chars, col_centers)
                    n_now = sum(1 for v in probe if v)
                if not n_now:
                    continue
                records = [
                    {"Date": report_date, "Properties": j + 1, "Additional Properties": j + 1, "Fluid set": "", "Source": ""}
                    for j in range(n_now)
                ]

            n_active = len(records)
            values = _bucket_chars_by_column(value_chars, col_centers[:n_active])

            if low_label in ("set de fluido", "fluid set"):
                for j in range(min(n_active, len(values))):
                    if values[j]:
                        records[j]["Fluid set"] = values[j]
                continue

            if low_label in ("origen", "source"):
                for j in range(min(n_active, len(values))):
                    if values[j]:
                        records[j]["Source"] = values[j]
                continue

            if low_label in ("tiempo", "time"):
                for j in range(min(n_active, len(values))):
                    time_txt = values[j]
                    if not time_txt:
                        continue
                    records[j]["Time"] = time_txt
                    ts = _mud_compose_datetime(report_date, time_txt)
                    records[j]["Date"] = ts if ts is not None else report_date
                    records[j]["DateTime"] = _mud_isoformat_no_tz(ts if ts is not None else report_date)
                continue

            for j in range(min(n_active, len(values))):
                raw_value = values[j]
                if not raw_value:
                    continue
                _mud_apply_daily_property(records[j], label, unit, raw_value)

    if not records:
        return []

    if not records[0].get("Source"):
        records[0]["Source"] = source_name

    return [
        r
        for r in records
        if any(k not in MUD_METADATA_COLUMNS and pd.notna(v) and _mud_clean_cell_text(v) not in ("", "/") for k, v in r.items())
    ]


# ============================================================
# Reporte diario Mi SWACO / Schlumberger ("PROPIEDADES DE LODO")
# ============================================================

# Etiqueta de fila normalizada -> campo interno. El orden importa: gana la primera
# expresión que matchea, así que las etiquetas específicas van antes que las genéricas
# ("solidos corregidos" antes de "solidos", "filtrado hthp" antes de "filtrado api").
_SWACO_ROW_FIELDS: list[tuple[str, str]] = [
    (r"^temp\.?\s*(flow\s*line|de\s*salida|salida)", "fl_temp"),
    (r"^profundidad\s*/\s*tvd", "depth"),
    (r"^densidad", "density"),
    (r"^visc\.?\s*embudo", "fv"),
    (r"^temp\.?\s*reolog", "rheo_temp"),
    (r"^r?\s*600\s*/\s*r?\s*300$", "l600_l300"),
    (r"^r?\s*200\s*/\s*r?\s*100$", "l200_l100"),
    (r"^r?\s*6\s*/\s*r?\s*3$", "l6_l3"),
    (r"^r\s*600$", "l600"),
    (r"^r\s*300$", "l300"),
    (r"^r\s*200$", "l200"),
    (r"^r\s*100$", "l100"),
    (r"^r\s*6$", "l6"),
    (r"^r\s*3$", "l3"),
    (r"^(vp|pv)$", "pv"),
    (r"^(pc|yp)$", "yp"),
    (r"^10\s*s\s*/\s*10\s*m\s*/\s*30\s*m\s*gel", "gel_triple"),
    (r"^gel\s*10\s*s", "gel_10s"),
    (r"^gel\s*10\s*m", "gel_10min"),
    (r"^gel\s*30\s*m", "gel_30min"),
    (r"^filtrado\s*(hthp|hpht)", "hthp"),
    (r"^filtrado\s*api", "api_filtrate"),
    (r"^revoque", "cake"),
    (r"^solidos\s+corregidos", "corr_solid"),
    (r"^solidos(\s+retorta)?$", "solids"),
    (r"^aceite$", "nap"),
    (r"^agua$", "water"),
    (r"^(rel\.?\s*)?aceite\s*/\s*agua$", "nap_ratio"),
    (r"^alcalinidad", "alkalinity"),
    (r"^cloruros\s+(en\s+)?lodo", "chlorides"),
    (r"^salinidad", "salinity"),
    (r"^cloruros\s+fase\s+acuosa", "water_phase_salinity"),
    (r"^ex(c)?\.?\s*(cal|calcio)", "excess_lime"),
    (r"^estabilidad\s+el", "electrical_stability"),
    (r"^arena$", "sand"),
    (r"^lgs(\s+totales)?$", "lgs_pct"),
]

# Secciones que arrancan justo debajo de la tabla de propiedades. Se evalúan sobre el
# texto recortado a la banda de la tabla, así que los paneles laterales del mismo
# renglón visual ("EQUIPOS DE CONTROL DE SOLIDOS", "PRODUCTOS USADOS") no cortan nada.
_SWACO_SECTION_STOP = (
    r"^(comentarios|distribucion|detalle|sumario|analisis|reologia|balance|"
    r"especificaciones|equipos|productos|volumenes|concentracion|tanque)\b"
)

_SWACO_LR_PAIRS = {
    "l600_l300": ("L600", "L300"),
    "l200_l100": ("L200", "L100"),
    "l6_l3": ("L6", "L3"),
}
_SWACO_SIMPLE_FIELDS = {
    "fl_temp": "FL Temp",
    "fv": "FV",
    "rheo_temp": "PV Temp",
    "l600": "L600",
    "l300": "L300",
    "l200": "L200",
    "l100": "L100",
    "l6": "L6",
    "l3": "L3",
    "pv": "PV",
    "yp": "YP",
    "gel_10s": "Gel_10s",
    "gel_10min": "Gel_10min",
    "gel_30min": "Gel_30min",
    "api_filtrate": "Filtrado",
    "corr_solid": "Corr Solid",
    "solids": "Solids",
    "nap": "NAP",
    "water": "Water",
    "alkalinity": "Alkalinity",
    "chlorides": "Chlorides",
    "salinity": "Salinity",
    "water_phase_salinity": "Water Phase Salinity",
    "excess_lime": "Excess Lime",
    "electrical_stability": "Electrical_Stability",
    "sand": "Sand",
    "lgs_pct": "LGS (%)",
}


def _swaco_normalize_label(text: str) -> str:
    return re.sub(r"\s+", " ", _mud_clean_cell_text(text)).strip(" :").lower()


def _swaco_row_field(label: str) -> str | None:
    low = _swaco_normalize_label(label)
    if not low:
        return None
    for pattern, field in _SWACO_ROW_FIELDS:
        if re.match(pattern, low):
            return field
    return None


def _swaco_group_words_into_rows(page, y_tolerance: float = 3.0) -> list[list[dict]]:
    words = sorted(page.extract_words(), key=lambda w: (w["top"], w["x0"]))
    rows: list[list[dict]] = []
    for w in words:
        if rows and abs(w["top"] - rows[-1][0]["top"]) <= y_tolerance:
            rows[-1].append(w)
        else:
            rows.append([w])
    for r in rows:
        r.sort(key=lambda w: w["x0"])
    return rows


def _swaco_sample_columns(row: list[dict], gap: float = 8.0) -> list[dict]:
    """
    Divide la fila 'Muestra / Hora' en columnas de muestra. Cada celda es del tipo
    'Succion / 20:00' o 'Succion - 12:00'; se exige que contenga una hora para no
    confundirlas con los encabezados de los paneles vecinos ('Productos', 'Tamaño',
    'Cantidad'), que están en el mismo renglón visual.
    """
    cells: list[list[dict]] = []
    for w in row:
        if cells and w["x0"] - cells[-1][-1]["x1"] <= gap:
            cells[-1].append(w)
        else:
            cells.append([w])

    cols: list[dict] = []
    for cell in cells:
        text = " ".join(w["text"] for w in cell)
        m = re.search(r"\d{1,2}:\d{2}", text)
        if not m:
            continue
        cols.append({
            "x0": min(w["x0"] for w in cell),
            "x1": max(w["x1"] for w in cell),
            "source": text[:m.start()].strip(" -/–"),
            "time": m.group(0),
        })
    return sorted(cols, key=lambda c: c["x0"])


def _swaco_split_label_unit(words: list[dict], big_gap: float = 12.0) -> tuple[str, str]:
    """Separa etiqueta y unidad por el hueco horizontal más grande: en esta plantilla
    la unidad va en su propia columna, muy a la derecha del texto de la etiqueta."""
    if not words:
        return "", ""
    best_i, best_gap = None, 0.0
    for i in range(1, len(words)):
        g = words[i]["x0"] - words[i - 1]["x1"]
        if g > best_gap:
            best_gap, best_i = g, i
    if best_i is not None and best_gap >= big_gap:
        label = " ".join(w["text"] for w in words[:best_i])
        unit = " ".join(w["text"] for w in words[best_i:])
    else:
        label, unit = " ".join(w["text"] for w in words), ""
    return label.strip(), unit.strip()


def _swaco_bucket_values(words: list[dict], centers: list[float]) -> list[str]:
    buckets: list[list[dict]] = [[] for _ in centers]
    for w in words:
        mid = (w["x0"] + w["x1"]) / 2
        j = min(range(len(centers)), key=lambda k: abs(centers[k] - mid))
        buckets[j].append(w)
    return [" ".join(w["text"] for w in sorted(b, key=lambda w: w["x0"])).strip() for b in buckets]


def _swaco_apply_field(rec: dict, field: str, raw: str) -> None:
    nums = _extract_all_numbers(raw)
    simple = _SWACO_SIMPLE_FIELDS.get(field)
    if simple:
        if nums:
            rec[simple] = nums[0]
        return
    if field == "depth":
        if nums:
            rec["Depth (MD)"] = nums[0]
        if len(nums) >= 2:
            rec["Depth (TVD)"] = nums[1]
        return
    if field == "density":
        if nums:
            rec["Density"] = nums[0]
            rec["Density @ °C"] = _mud_pair_string(raw)
        if len(nums) >= 2:
            rec["Density Temp"] = nums[1]
        return
    if field in _SWACO_LR_PAIRS:
        hi, lo = _SWACO_LR_PAIRS[field]
        if nums:
            rec[hi] = nums[0]
        if len(nums) >= 2:
            rec[lo] = nums[1]
        return
    if field == "gel_triple":
        g1, g2, g3 = _parse_gel_triple(raw)
        if g1 is not None:
            rec["Gel_10s"] = g1
        if g2 is not None:
            rec["Gel_10min"] = g2
        if g3 is not None:
            rec["Gel_30min"] = g3
        return
    if field == "hthp":
        if nums:
            rec["HTHP"] = nums[0]
        if len(nums) >= 2:
            rec["HTHP @ °C"] = nums[1]
        return
    if field == "cake":
        # 'Revoque API / HTHP' viene como '/ 1' (API vacío, HTHP = 1) o '1 / 1'.
        if nums:
            rec["Cake (HTHP)"] = nums[-1]
        return
    if field == "nap_ratio":
        if nums:
            rec["NAP Ratio"] = nums[0]
            rec["NAP 2"] = nums[0]
        if len(nums) >= 2:
            rec["Water Ratio"] = nums[1]
        return


# Panel 'ANALISIS DE SOLIDOS'. Las dos ultimas entradas no se exportan: sirven para
# identificar a que muestra pertenece el panel (ver _swaco_attach_solids_panel).
_SWACO_SOLIDS_PANEL = [
    ("ASG", r"Average\s+SG\s+Solids\s+([\d.,]+)"),
    ("HGS (kg/m³)", r"High\s+Gravity\s+Wt\.?\s+kg/m\S*\s+([\d.,]+)"),
    ("LGS (kg/m³)", r"Low\s+Gravity\s+Wt\.?\s+kg/m\S*\s+([\d.,]+)"),
    ("HGS (%)", r"High\s+Gravity\s+%\s+([\d.,]+)"),
    ("LGS (%)", r"Low\s+Gravity\s+%\s+([\d.,]+)"),
    ("Corr Solid", r"Adjusted\s+Solids\s+%vol\s+([\d.,]+)"),
    ("Salinity", r"Salt\s+Wt\s+%wt\s+([\d.,]+)"),
    ("NAP Ratio", r"Oil/Water\s+Ratio\s+([\d.,]+)"),
]
_SWACO_SOLIDS_EXPORTED = ("ASG", "HGS (kg/m³)", "LGS (kg/m³)", "HGS (%)", "LGS (%)", "Corr Solid")


def _swaco_solids_panel(full_text: str) -> dict:
    """Lee el panel 'ANALISIS DE SOLIDOS' del reporte (mismo texto en PDF y Excel)."""
    panel: dict = {}
    for key, pattern in _SWACO_SOLIDS_PANEL:
        m = re.search(pattern, full_text, re.IGNORECASE)
        if not m:
            continue
        val = _extract_numeric(m.group(1))
        if val is not None:
            panel[key] = val
    return panel


def _swaco_attach_solids_panel(records: list[dict], panel: dict) -> None:
    """
    El panel describe UNA sola muestra —la del cierre del reporte—, no un promedio del
    dia. En el N29 'Adjusted Solids' 21,24 y 'Low Gravity %' 0,02 coinciden exactamente
    con la muestra de las 20:00 y no con las de 14:00 y 04:00 (23,24 y 1,9). Se localiza
    esa muestra comparando los campos que el panel repite de la tabla de propiedades, y
    solo ahi se escriben ASG/HGS/LGS: repartirlos a todas las filas sembraba de valores
    ajenos las demas muestras.
    """
    if not records or not panel:
        return
    best_idx, best_score = None, 0
    for i, rec in enumerate(records):
        score = 0
        for key in ("Corr Solid", "Salinity", "LGS (%)", "NAP Ratio"):
            pv, rv = panel.get(key), rec.get(key)
            if pv is None or rv is None:
                continue
            if abs(float(pv) - float(rv)) <= max(0.05, abs(float(pv)) * 0.01):
                score += 1
        if score > best_score:
            best_idx, best_score = i, score
    # Sin coincidencias, la primera columna del reporte es la del cierre.
    target = records[best_idx if best_idx is not None else 0]
    for key in _SWACO_SOLIDS_EXPORTED:
        if panel.get(key) is not None:
            target.setdefault(key, panel[key])


def _swaco_report_header(full_text: str) -> tuple[pd.Timestamp | None, str, str]:
    """Fecha, tipo de lodo y pozo del encabezado. Sirve para PDF y para Excel."""
    report_date = None
    # 'Fecha :' y no 'Fecha de Incio :'. El Excel trae un datetime real ("2026-08-12
    # 00:00:00") y el PDF un dd/mm/aaaa; _extract_date_from_text acepta los dos.
    m = re.search(r"\bFecha\s*:\s*([^\n]{0,40})", full_text, re.IGNORECASE)
    if m:
        report_date = _extract_date_from_text(m.group(1))
    m_fluid = re.search(r"Tipo\s+de\s+Lodo\s*:?\s*(\S+)", full_text, re.IGNORECASE)
    m_well = re.search(r"\bPozo\s*:?\s*(\S+)", full_text, re.IGNORECASE)
    return (
        report_date,
        m_fluid.group(1).strip(":") if m_fluid else "",
        m_well.group(1).strip(":") if m_well else "",
    )


def _parse_mud_pdf_swaco_daily(pdf, source_name: str = "") -> list[dict]:
    """
    Parser posicional del reporte diario de fluidos Mi SWACO / Schlumberger
    ('DRILLING SOLUTIONS', sección 'PROPIEDADES DE LODO'). Soporta tanto la versión
    con varias muestras en columnas (Succion 20:00 / 14:00 / 4:00) como la de una
    sola columna, y devuelve un registro por muestra con las claves canónicas de la
    bitácora, igual que _parse_mud_pdf_daily_report_grid con el layout WellSight.
    """
    full_text = "\n".join(page.extract_text() or "" for page in pdf.pages)
    if not re.search(r"propiedades\s+de\s+lodo", full_text, re.IGNORECASE):
        return []
    if not re.search(r"muestra\s*/\s*hora", full_text, re.IGNORECASE):
        return []

    # El pozo no va a la bitácora exportada (no es columna del formato), pero identifica
    # la muestra al acumular varios reportes: evita fusionar dos pozos que coincidan en
    # fecha, hora y origen.
    report_date, fluid_set, well = _swaco_report_header(full_text)
    report_date = report_date or _date_from_filename_or_today(source_name)

    records: list[dict] = []
    for page in pdf.pages:
        rows = _swaco_group_words_into_rows(page)
        header_idx, cols = None, []
        for i, row in enumerate(rows):
            if not re.match(r"muestra\s*/\s*hora", " ".join(w["text"] for w in row), re.IGNORECASE):
                continue
            found = _swaco_sample_columns(row)
            if found:
                header_idx, cols = i, found
                break
        if header_idx is None:
            continue

        anchor_x0 = rows[header_idx][0]["x0"]
        centers = [(c["x0"] + c["x1"]) / 2 for c in cols]
        if len(centers) > 1:
            half = min(centers[k + 1] - centers[k] for k in range(len(centers) - 1)) / 2
        else:
            half = max(20.0, (cols[0]["x1"] - cols[0]["x0"]) / 2 + 5)
        band_left, band_right = centers[0] - half, centers[-1] + half

        # Las muestras vienen en el orden del reporte (la más reciente primero), pero la
        # bitácora numera 'Properties' en orden cronológico como el layout WellSight.
        order = sorted(range(len(cols)), key=lambda j: _mud_parse_time_value(cols[j]["time"]) or datetime.min.time())
        rank = {j: k + 1 for k, j in enumerate(order)}

        page_records: list[dict] = []
        for j, c in enumerate(cols):
            t = _mud_parse_time_value(c["time"])
            time_txt = t.strftime("%H:%M") if t else c["time"]
            ts = _mud_compose_datetime(report_date, time_txt) or report_date
            page_records.append({
                "Date": ts,
                "DateTime": _mud_isoformat_no_tz(ts),
                "Time": time_txt,
                "Properties": rank[j],
                "Additional Properties": rank[j],
                "Fluid set": fluid_set,
                "Source": c["source"] or source_name,
                "Well": well,
            })

        for row in rows[header_idx + 1:]:
            band_text = " ".join(w["text"] for w in row if w["x1"] <= band_right + 4)
            if re.match(_SWACO_SECTION_STOP, _swaco_normalize_label(band_text)):
                break
            # Toda etiqueta de propiedad arranca en la misma x que 'Muestra'. Filtrarlo
            # descarta los paneles a la derecha ('ESPECIFICACIONES DE LODO', que repite
            # Densidad/VP/Filtrado con los valores del cierre del día).
            if row[0]["x0"] > anchor_x0 + 8:
                continue

            label, _unit = _swaco_split_label_unit([w for w in row if w["x1"] <= band_left])
            field = _swaco_row_field(label)
            if not field:
                continue
            values = _swaco_bucket_values(
                [w for w in row if w["x0"] >= band_left and w["x1"] <= band_right + 4], centers
            )
            for j, rec in enumerate(page_records):
                if j < len(values) and values[j]:
                    _swaco_apply_field(rec, field, values[j])

        records.extend(page_records)

    if not records:
        return []

    _swaco_attach_solids_panel(records, _swaco_solids_panel(full_text))
    return _swaco_finalize_records(records)


def _swaco_finalize_records(records: list[dict]) -> list[dict]:
    for rec in records:
        # El reporte no da temperatura propia para la viscosidad de embudo: se mide sobre
        # la misma muestra que la densidad, y así queda 'Fv @ Nº°C' alineada con 'D @ Nº°C'
        # como en la bitácora del layout WellSight.
        if "FV" in rec and "FV Temp" not in rec and "Density Temp" in rec:
            rec["FV Temp"] = rec["Density Temp"]

    return [
        r
        for r in records
        if any(
            k not in MUD_METADATA_COLUMNS and pd.notna(v) and _mud_clean_cell_text(v) not in ("", "/")
            for k, v in r.items()
        )
    ]


def _parse_mud_swaco_sheet(df_raw: pd.DataFrame, source_name: str = "") -> list[dict]:
    """
    El mismo reporte diario Mi SWACO pero en Excel: la plantilla que genera el PDF.
    Estructura fija de la hoja 'OBM': columna A etiqueta, columna B unidad y una columna
    por muestra a partir de la C. Reutiliza el mapa de etiquetas del parser de PDF, así
    que ambos formatos rinden exactamente la misma bitácora.
    """
    if df_raw is None or df_raw.empty or df_raw.shape[0] < 10:
        return []

    def cell(i: int, j: int) -> str:
        if i >= df_raw.shape[0] or j >= df_raw.shape[1]:
            return ""
        return _mud_clean_cell_text(df_raw.iat[i, j])

    # Texto por filas: reproduce el renglón del PDF, así el encabezado y el panel
    # 'ANALISIS DE SOLIDOS' se leen con las mismas expresiones que en el PDF.
    row_text = [
        " ".join(t for t in (cell(i, j) for j in range(df_raw.shape[1])) if t)
        for i in range(df_raw.shape[0])
    ]
    full_text = "\n".join(row_text)
    if not re.search(r"propiedades\s+de\s+lodo", full_text, re.IGNORECASE):
        return []

    hdr = None
    for i in range(df_raw.shape[0]):
        if re.match(r"muestra\s*/\s*hora", _swaco_normalize_label(cell(i, 0))):
            hdr = i
            break
    if hdr is None:
        return []

    sample_cols: list[tuple[int, str, str]] = []
    for j in range(1, df_raw.shape[1]):
        txt = cell(hdr, j)
        m = re.search(r"\d{1,2}:\d{2}", txt)
        if m:
            sample_cols.append((j, txt[:m.start()].strip(" -/–"), m.group(0)))
    if not sample_cols:
        return []

    report_date, fluid_set, well = _swaco_report_header(full_text)
    report_date = report_date or _date_from_filename_or_today(source_name)

    order = sorted(range(len(sample_cols)), key=lambda k: _mud_parse_time_value(sample_cols[k][2]) or datetime.min.time())
    rank = {k: n + 1 for n, k in enumerate(order)}

    records: list[dict] = []
    for k, (_j, src, time_raw) in enumerate(sample_cols):
        t = _mud_parse_time_value(time_raw)
        time_txt = t.strftime("%H:%M") if t else time_raw
        ts = _mud_compose_datetime(report_date, time_txt) or report_date
        records.append({
            "Date": ts,
            "DateTime": _mud_isoformat_no_tz(ts),
            "Time": time_txt,
            "Properties": rank[k],
            "Additional Properties": rank[k],
            "Fluid set": fluid_set,
            "Source": src or source_name,
            "Well": well,
        })

    for i in range(hdr + 1, df_raw.shape[0]):
        label = cell(i, 0)
        if re.match(_SWACO_SECTION_STOP, _swaco_normalize_label(label)):
            break
        field = _swaco_row_field(label)
        if not field:
            continue
        for k, (j, _src, _t) in enumerate(sample_cols):
            raw = cell(i, j)
            # La plantilla rellena las muestras vacías con 0 y con espacios; tomarlos
            # como medición dejaría ceros falsos en la bitácora.
            if not raw or raw == "0":
                continue
            _swaco_apply_field(records[k], field, raw)

    _swaco_attach_solids_panel(records, _swaco_solids_panel(full_text))
    return _swaco_finalize_records(records)


def _parse_mud_pdf(file, source_name: str = "") -> list[dict]:
    """Extrae tablas/texto de PDF y parsea propiedades conocidas."""
    out: list[dict] = []
    try:
        import pdfplumber  # type: ignore
    except ImportError:
        return out
    name = source_name or getattr(file, "name", "") or ""
    row_record: dict = {"Date": _date_from_filename_or_today(name), "Source": name}
    try:
        with pdfplumber.open(file) as pdf:
            grid_rows = _parse_mud_pdf_daily_report_grid(pdf, name)
            if grid_rows:
                return grid_rows

            swaco_rows = _parse_mud_pdf_swaco_daily(pdf, name)
            if swaco_rows:
                return swaco_rows

            full_text_parts = []
            for page in pdf.pages:
                page_text = page.extract_text() or ""
                if page_text:
                    full_text_parts.append(page_text)
                tables = page.extract_tables() or []
                for table in tables:
                    for row in table or []:
                        if not row:
                            continue
                        for idx, cell in enumerate(row):
                            if cell is None:
                                continue
                            canonical = _normalize_mud_property_name(str(cell))
                            if not canonical:
                                continue
                            for other in row[idx + 1:]:
                                if other is None:
                                    continue
                                _mud_apply_canonical_value(row_record, canonical, other)
                                if canonical in row_record or canonical.startswith("Gel"):
                                    break
                        full_text = "\n".join(full_text_parts)
            dt_pdf = _extract_date_from_text(full_text)
            if dt_pdf is not None:
                row_record["Date"] = dt_pdf
            _parse_mud_text_block(full_text, row_record)
            _parse_mud_lines(full_text, row_record)
        if any(k for k in row_record if k not in ("Date", "Source")):
            out.append(row_record)
    except Exception:
        pass
    return out


def _fetch_mud_attachments_from_email(
    imap_server: str,
    imap_user: str,
    imap_pass: str,
    filename_contains: str | None = None,
    mark_read: bool = True,
) -> list[tuple[str, bytes]]:
    """
    Descarga adjuntos PDF/Excel/CSV de correos no leídos por IMAP.
    filename_contains: filtro opcional (ej. "Daily Full Report" o "LA-358").
    mark_read: si True, marca los correos como leídos tras descargar.
    Retorna lista de (nombre_archivo, contenido_bytes).
    """
    results: list[tuple[str, bytes]] = []
    try:
        import imaplib
        import email as email_module
    except ImportError:
        return results
    try:
        with imaplib.IMAP4_SSL(imap_server, timeout=30) as mail:
            mail.login(imap_user, imap_pass)
            mail.select("inbox")
            status, messages = mail.search(None, "(UNSEEN)")
            if status != "OK":
                return results
            for num in (messages[0] or b"").split():
                if not num:
                    continue
                status, data = mail.fetch(num, "(RFC822)")
                if status != "OK":
                    continue
                msg = email_module.message_from_bytes(data[0][1])
                for part in msg.walk():
                    if part.get_content_disposition() != "attachment":
                        continue
                    filename = part.get_filename()
                    if not filename:
                        continue
                    filename = str(filename).strip()
                    ext = (filename or "").lower()
                    if not (
                        ext.endswith(".pdf")
                        or ext.endswith(".xlsx")
                        or ext.endswith(".xls")
                        or ext.endswith(".csv")
                    ):
                        continue
                    if filename_contains and filename_contains.strip():
                        if filename_contains.strip().lower() not in filename.lower():
                            continue
                    payload = part.get_payload(decode=True)
                    if payload:
                        results.append((filename, bytes(payload)))
                if mark_read and results:
                    try:
                        mail.store(num, "+FLAGS", "\\Seen")
                    except Exception:
                        pass
    except Exception:
        raise
    return results


def _build_mud_bitacora(parsed_rows: list[dict]) -> pd.DataFrame:
    """Construye DataFrame bitácora con columnas canónicas."""
    if not parsed_rows:
        return pd.DataFrame()
    all_keys = set()
    for r in parsed_rows:
        all_keys.update(r.keys())
    cols = [c for c in MUD_CANONICAL_ORDER if c in all_keys]
    for c in sorted(all_keys):
        if c not in cols:
            cols.append(c)
    rows = []
    for r in parsed_rows:
        row = {}
        for k in cols:
            row[k] = r.get(k)
        rows.append(row)
    df = pd.DataFrame(rows)
    if "Date" in df.columns:
        df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
        df = df.dropna(subset=["Date"]).sort_values("Date").reset_index(drop=True)
    if "DateTime" in df.columns:
        df["DateTime"] = df["DateTime"].fillna("")
    return df


def _mud_parse_attachment(name: str, data: bytes) -> list[dict]:
    """Parsea un reporte (subido o descargado del correo) según su extensión."""
    low = (name or "").lower()
    if low.endswith(".pdf"):
        return _parse_mud_pdf(io.BytesIO(data), source_name=name)
    if low.endswith((".xlsx", ".xls")):
        out: list[dict] = []
        xl = pd.ExcelFile(io.BytesIO(data))
        sheets = [(sh, pd.read_excel(xl, sheet_name=sh, header=None)) for sh in xl.sheet_names[:5]]
        # El reporte Mi SWACO en Excel trae otras hojas (volúmenes, inventario) que el
        # parser genérico leería como propiedades: si una hoja es el reporte, se devuelve
        # esa y no se sigue.
        for _sh, df_raw in sheets:
            swaco_rows = _parse_mud_swaco_sheet(df_raw, name)
            if swaco_rows:
                return swaco_rows
        for _sh, df_raw in sheets:
            out.extend(_parse_mud_excel_sheet(df_raw, name))
        return out
    df_raw = pd.read_csv(io.BytesIO(data), sep=None, engine="python", low_memory=False)
    return _parse_mud_csv(df_raw, name)


def _mud_file_signature(data: bytes) -> str:
    """Huella del contenido: permite no re-parsear en cada rerun de Streamlit y no
    duplicar el mismo reporte aunque llegue con otro nombre."""
    return f"{hashlib.md5(data).hexdigest()}:{len(data)}"


def _mud_renumber_properties(df: pd.DataFrame) -> pd.DataFrame:
    """Renumera 'Properties' por día. En una bitácora acumulada cada jornada vuelve a
    contar desde 1, que es el significado original de la columna (N° de muestra del día)
    y deja el resultado idéntico al de antes cuando hay un solo día."""
    if df.empty or "Date" not in df.columns:
        return df
    df = df.copy()
    seq = (df.groupby(df["Date"].dt.normalize(), sort=False).cumcount() + 1).values
    df["Properties"] = seq
    df["Additional Properties"] = seq
    return df


def _mud_merge_bitacora(existing: pd.DataFrame | None, new: pd.DataFrame | None) -> pd.DataFrame:
    """
    Funde dos bitácoras en una sola ordenada por fecha y hora. Una muestra queda
    identificada por fecha/hora + origen + pozo: así el reporte de las 20 h, que repite
    las muestras de las 04 h y 14 h del mismo día, actualiza esas filas en vez de
    duplicarlas (gana la última ingestada). Vuelve a numerar 'Properties' por día.
    """
    frames = [f for f in (existing, new) if f is not None and not f.empty]
    if not frames:
        return pd.DataFrame()
    df = pd.concat(frames, ignore_index=True, sort=False)
    if "Date" not in df.columns:
        return df
    df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
    df = df.dropna(subset=["Date"])
    if df.empty:
        return df

    def _txt(col: str) -> pd.Series:
        if col not in df.columns:
            return pd.Series("", index=df.index)
        return df[col].astype(str).str.strip().str.lower()

    key = pd.DataFrame({
        "d": df["Date"].dt.strftime("%Y-%m-%dT%H:%M"),
        "s": _txt("Source"),
        "w": _txt("Well"),
    })
    df = df[~key.duplicated(keep="last")]
    df = df.sort_values("Date", kind="stable").reset_index(drop=True)
    return _mud_renumber_properties(df)


def _mud_numeric_property_columns(bitacora: pd.DataFrame) -> list[str]:
    preferred = [
        "Density", "FL Temp", "Density @ °C", "FV", "FV Temp", "PV", "PV Temp", "YP",
        "Gel_10s", "Gel_10min", "Gel_30min", "tau0",
        "L600", "L300", "L200", "L100", "L6", "L3",
        "HTHP", "HTHP @ °C", "Corr Solid", "NAP", "Water", "NAP Ratio", "Water Ratio",
        "Sand", "Cake (HTHP)", "Chlorides", "Calcium", "CaCl2", "Water Phase Salinity",
        "Excess Lime", "Electrical_Stability", "LGS (%)", "HGS (%)", "LGS (kg/m³)", "HGS (kg/m³)",
        "ASG", "n (HB)", "K (HB)", "Viscometer Sag Shoe Test", "(VSST)",
        "Marsh", "Temperature", "VA", "Filtrado", "Enjarre", "LGS", "HGS", "Solids", "Oil", "RAA",
        "AgNO3", "Salinity", "Alkalinity", "Excess_Cal",
    ]
    cols = []
    for c in preferred:
        if c in bitacora.columns and pd.api.types.is_numeric_dtype(bitacora[c]):
            cols.append(c)
    for c in bitacora.columns:
        if c in cols or c == "Date" or c in MUD_ANALYTIC_EXCLUDE:
            continue
        if pd.api.types.is_numeric_dtype(bitacora[c]):
            cols.append(c)
    return cols


def _mud_effective_value_temp(row, value_col: str, temp_col: str, pair_col: str):
    """Valor y temperatura efectivos de una fila, recurriendo al par 'valor @ temp' si falta el numérico."""
    val = row.get(value_col)
    temp = row.get(temp_col)
    if pd.isna(val) and pd.notna(row.get(pair_col)):
        nums = _extract_all_numbers(row.get(pair_col))
        if nums:
            val = nums[0]
        if len(nums) >= 2:
            temp = nums[1]
    return val, temp


def _mud_dynamic_temp_spread(view: pd.DataFrame, value_col: str, temp_col: str, pair_col: str, label_prefix: str, max_cols: int = 8) -> list[str]:
    """
    Reparte cada fila en una columna '{label_prefix} @ {temp}°C' según SU temperatura
    real medida (no fuerza a 54/45/44°C, que solo aplican cuando el reporte mide justo
    a esas tres). El orden de las columnas es el de aparición en la bitácora, así que
    para un reporte que sí mide a 54/45/44°C el resultado es idéntico al de antes.
    """
    col_by_temp: dict[float, str] = {}
    for _, row in view.iterrows():
        val, temp = _mud_effective_value_temp(row, value_col, temp_col, pair_col)
        if pd.isna(val) or pd.isna(temp):
            continue
        key = round(float(temp), 1)
        if key not in col_by_temp and len(col_by_temp) < max_cols:
            col_by_temp[key] = f"{label_prefix} @ {_mud_num_to_text(key)}°C"

    for col in col_by_temp.values():
        if col not in view.columns:
            view[col] = np.nan

    for idx, row in view.iterrows():
        val, temp = _mud_effective_value_temp(row, value_col, temp_col, pair_col)
        if pd.isna(val) or pd.isna(temp):
            continue
        col = col_by_temp.get(round(float(temp), 1))
        if col:
            view.at[idx, col] = _extract_numeric(val)

    return list(col_by_temp.values())


def _mud_spread_temp_columns(view: pd.DataFrame) -> pd.DataFrame:
    """Distribuye densidad y FV según la temperatura real de cada muestra, y PV/HTHP en sus columnas fijas (siempre 65°C/149°C por procedimiento)."""
    _mud_dynamic_temp_spread(view, "Density", "Density Temp", "Density @ °C", "D")
    _mud_dynamic_temp_spread(view, "FV", "FV Temp", "FV @ °C", "Fv")

    for col in ["PV @ 65°C", "HTHP @ 149°C", "NAP 2"]:
        if col not in view.columns:
            view[col] = np.nan

    for idx, row in view.iterrows():
        pv_val = row.get("PV")
        if pd.notna(pv_val):
            view.at[idx, "PV @ 65°C"] = pv_val

        hthp_val = row.get("HTHP")
        if pd.notna(hthp_val):
            view.at[idx, "HTHP @ 149°C"] = hthp_val

        nap_ratio = row.get("NAP Ratio")
        if pd.notna(nap_ratio):
            view.at[idx, "NAP 2"] = nap_ratio

    return view


_MUD_FIXED_DENSITY_COLS = {"D @ 54°C", "D @ 45°C", "D @ 44°C"}
_MUD_FIXED_FV_COLS = {"Fv @ 54°C", "Fv @ 45°C", "Fv @ 44°C"}


def _mud_export_specs_for(view: pd.DataFrame) -> list[tuple[str, str, str]]:
    """
    Mismo orden que MUD_EXPORT_HEADER_SPECS, pero sustituyendo las 3 columnas fijas de
    densidad/Fv (54/45/44°C) por las columnas D/Fv @ {temp}°C que realmente existan en
    esta bitácora (según la temperatura real de cada muestra), preservando el orden en
    que _mud_dynamic_temp_spread las agregó.
    """
    density_dynamic = [c for c in view.columns if re.fullmatch(r"D @ [\d.]+°C", c)]
    fv_dynamic = [c for c in view.columns if re.fullmatch(r"Fv @ [\d.]+°C", c)]

    specs: list[tuple[str, str, str]] = []
    density_done = fv_done = False
    for col_name, h1, h2 in MUD_EXPORT_HEADER_SPECS:
        if col_name in _MUD_FIXED_DENSITY_COLS:
            if not density_done:
                specs.extend((c, c, "kg/m³") for c in density_dynamic)
                density_done = True
            continue
        if col_name in _MUD_FIXED_FV_COLS:
            if not fv_done:
                specs.extend((c, c, "s/qt") for c in fv_dynamic)
                fv_done = True
            continue
        # Solo las columnas que el reporte realmente trae. Calcio, CaCl2, Tauy, n, k o el
        # ensayo de asentamiento no existen en el reporte Mi SWACO, y dejarlas en blanco
        # se lee como «se midió y salió vacío», que no es lo mismo que «no se mide».
        if col_name not in view.columns:
            continue
        specs.append((col_name, h1, h2))
    return specs


def _mud_build_view_df(bitacora: pd.DataFrame) -> pd.DataFrame:
    if bitacora is None or bitacora.empty:
        return pd.DataFrame()
    view = bitacora.copy()
    if "DateTime" not in view.columns and "Date" in view.columns:
        view["DateTime"] = pd.to_datetime(view["Date"], errors="coerce").dt.strftime("%Y-%m-%dT%H:%M:%S")
    else:
        dt_series = pd.to_datetime(view.get("Date"), errors="coerce")
        mask = view["DateTime"].astype(str).str.strip().eq("")
        if mask.any():
            view.loc[mask, "DateTime"] = dt_series.loc[mask].dt.strftime("%Y-%m-%dT%H:%M:%S")
    if "Time" not in view.columns and "Date" in view.columns:
        view["Time"] = pd.to_datetime(view["Date"], errors="coerce").dt.strftime("%H:%M")
    if "Properties" not in view.columns:
        view["Properties"] = np.arange(1, len(view) + 1)
    if "Additional Properties" not in view.columns:
        view["Additional Properties"] = view["Properties"]
    view = _mud_spread_temp_columns(view)
    export_cols = [c for c, _, _ in _mud_export_specs_for(view)]
    for c in export_cols:
        if c not in view.columns:
            view[c] = np.nan
    for c in export_cols:
        if c in ("DateTime", "Time", "Fluid set", "Source"):
            view[c] = view[c].fillna("")
            continue
        # Los huecos llegan como None (propiedad ausente en ese reporte) y en una columna
        # de tipo object la tabla los pinta como el texto "None". Al acumular reportes de
        # distinta cobertura eso es la mayoría de la hoja, así que se pasan a numérico —
        # solo si ningún valor real se pierde en la conversión.
        num = pd.to_numeric(view[c], errors="coerce")
        if num.notna().sum() >= view[c].notna().sum():
            view[c] = num

    # Y fuera también las que existen pero quedaron sin un solo valor: p.ej. 'Revoque
    # API / HTHP', que es una fila del reporte pero puede venir en blanco. Se conservan
    # siempre las que dan identidad a la muestra, para que la hoja no pierda el esqueleto.
    keep_always = {"Properties", "Time", "DateTime"}
    cols_con_dato = [
        c for c in export_cols
        if c in keep_always or view[c].notna().any() and view[c].astype(str).str.strip().ne("").any()
    ]

    # 'Additional Properties' tampoco existe en el reporte Mi SWACO: es el N° de muestra
    # repetido. Viene del formato WellSight, que numera dos bloques ('Propiedades' y
    # 'Propiedades Adicionales'). Si no dice nada distinto de 'Properties', se va.
    if "Additional Properties" in cols_con_dato and "Properties" in view.columns:
        if view["Additional Properties"].astype(str).equals(view["Properties"].astype(str)):
            cols_con_dato.remove("Additional Properties")

    return view[cols_con_dato]


def _mud_display_df(view_df: pd.DataFrame) -> pd.DataFrame:
    """
    Copia solo para mostrar en pantalla: st.dataframe rotula los NaN como el texto
    "None", y en una bitácora acumulada —donde cada reporte trae distintas propiedades y
    distintas temperaturas de medición— eso llena la hoja de "None". Los exports siguen
    usando el DataFrame numérico.
    """
    if view_df is None or view_df.empty:
        return view_df
    return view_df.astype(object).where(view_df.notna(), "")


def _export_mud_bitacora_excel(view_df: pd.DataFrame, lang: str = MUD_LANG_ES) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Mud Bitacora Parser"

    headers = _mud_localize_specs(_mud_export_specs_for(view_df), lang)
    last_col = len(headers)

    title_date = _mud_bitacora_title_date(view_df)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    title_cell = ws.cell(1, 1)
    title_cell.value = _mud_bitacora_title(lang, title_date)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    title_cell.font = Font(bold=True)
    title_cell.fill = PatternFill("solid", fgColor="F2F2F2")
    ws.row_dimensions[1].height = 30

    fill_header = PatternFill("solid", fgColor="BFBFBF")
    fill_sub = PatternFill("solid", fgColor="E6E6E6")
    thin_gray = Side(style="thin", color="BFBFBF")
    border = Border(top=thin_gray, bottom=thin_gray)

    for idx, (_, h1, h2) in enumerate(headers, start=1):
        c2 = ws.cell(2, idx, h1)
        c3 = ws.cell(3, idx, h2)
        for cell in (c2, c3):
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.font = Font(bold=True)
            cell.border = border
        c2.fill = fill_header
        c3.fill = fill_sub

    # Altura explícita en las dos filas de encabezado: con wrap_text y sin altura, Excel
    # no siempre reajusta la fila al abrir y recorta el texto que no cabe, así que un
    # encabezado de dos líneas ('Solidos Alta Gravedad') puede verse en blanco.
    ws.row_dimensions[2].height = 32
    ws.row_dimensions[3].height = 18

    widths = {
        "Depth (MD)": 13.71, "Depth (TVD)": 14.71, "Properties": 13.71, "Fluid set": 12.71,
        "Source": 10.71, "Time": 13.0, "DateTime": 20.71, "FL Temp": 10.71,
        "D @ 54°C": 11.71, "D @ 45°C": 13.0, "D @ 44°C": 13.0,
        "Fv @ 54°C": 12.71, "Fv @ 45°C": 13.0, "Fv @ 44°C": 13.0,
        "PV @ 65°C": 13.0, "YP": 10.71, "Gel_10s": 11.71, "Gel_10min": 13.71,
        "Gel_30min": 13.0, "tau0": 10.71, "L600": 14.71, "L300": 13.0, "L200": 13.0,
        "L100": 13.0, "L6": 12.71, "L3": 13.0, "HTHP @ 149°C": 15.71,
        "Corr Solid": 13.71, "NAP": 10.71, "Water": 13.0, "NAP 2": 13.0,
        "Water Ratio": 14.71, "Sand": 10.71, "Cake (HTHP)": 14.71, "Chlorides": 12.71,
        "Calcium": 10.71, "CaCl2": 13.0, "Water Phase Salinity": 22.71, "Excess Lime": 14.71,
        "Electrical_Stability": 18.71, "LGS (%)": 10.71, "HGS (%)": 13.0,
        "LGS (kg/m³)": 13.0, "HGS (kg/m³)": 13.0, "ASG": 13.0,
        "Additional Properties": 22.71, "n (HB)": 10.71, "K (HB)": 13.0,
        "Viscometer Sag Shoe Test": 22.71,
    }
    num_format = "0.00"
    int_format = "0"
    row_start = 4
    for r_idx, (_, row) in enumerate(view_df.iterrows(), start=row_start):
        for c_idx, (col_name, _, _) in enumerate(headers, start=1):
            val = row.get(col_name)
            cell = ws.cell(r_idx, c_idx, val)
            if pd.isna(val):
                cell.value = None
            elif col_name == "DateTime" and str(val).strip():
                cell.number_format = "@"
            elif isinstance(val, (int, np.integer)):
                cell.number_format = int_format
            elif isinstance(val, (float, np.floating)) and np.isfinite(float(val)):
                cell.number_format = num_format if abs(float(val) - round(float(val))) > 1e-9 else int_format
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[r_idx].height = 21

    for idx, (col_name, _, _) in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = widths.get(col_name, 12)
    ws.freeze_panes = "A4"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _mud_pdf_escape(text) -> str:
    s = _mud_clean_cell_text(text)
    return s.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def _mud_bitacora_title_date(view_df: pd.DataFrame) -> str:
    if "Date" in view_df.columns:
        dt0 = pd.to_datetime(view_df["Date"], errors="coerce")
        if hasattr(dt0, "notna") and dt0.notna().any():
            return dt0.dropna().min().strftime("%Y-%m-%d")
    if "DateTime" in view_df.columns:
        dt0 = pd.to_datetime(view_df["DateTime"], errors="coerce")
        if hasattr(dt0, "notna") and dt0.notna().any():
            return dt0.dropna().min().strftime("%Y-%m-%d")
    return ""


def _mud_pdf_format_cell(val) -> str:
    if val is None or (isinstance(val, float) and (pd.isna(val) or not np.isfinite(val))):
        return ""
    if isinstance(val, (int, np.integer)):
        return str(int(val))
    if isinstance(val, (float, np.floating)):
        return _mud_num_to_text(val)
    return _mud_clean_cell_text(val)


def _export_mud_bitacora_pdf(view_df: pd.DataFrame, lang: str = MUD_LANG_ES) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER
    from reportlab.lib.pagesizes import landscape, letter
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    if view_df is None or view_df.empty:
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
        doc.build([Paragraph("Sin datos para exportar.", getSampleStyleSheet()["Normal"])])
        return buffer.getvalue()

    headers = _mud_localize_specs(_mud_export_specs_for(view_df), lang)
    title_date = _mud_bitacora_title_date(view_df)
    title_text = _mud_bitacora_title(lang, title_date)

    buffer = io.BytesIO()
    page_size = landscape(letter)
    doc = SimpleDocTemplate(
        buffer,
        pagesize=page_size,
        rightMargin=6 * mm,
        leftMargin=6 * mm,
        topMargin=8 * mm,
        bottomMargin=8 * mm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "mud_pdf_title",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=11,
        alignment=TA_CENTER,
        leading=13,
    )
    hdr_style = ParagraphStyle(
        "mud_pdf_hdr",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=5,
        alignment=TA_CENTER,
        leading=6,
    )
    cell_style = ParagraphStyle(
        "mud_pdf_cell",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=4.5,
        alignment=TA_CENTER,
        leading=5,
    )

    green = colors.HexColor("#68cbb3")
    border = colors.HexColor("#222222")
    hdr_bg = colors.HexColor("#D9D9D9")
    sub_bg = colors.HexColor("#EDEDED")

    story = []
    title_tbl = Table([[Paragraph(_mud_pdf_escape(title_text), title_style)]], colWidths=[page_size[0] - 12 * mm])
    title_tbl.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.5, border),
        ("BACKGROUND", (0, 0), (-1, -1), green),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(title_tbl)
    story.append(Spacer(1, 4))

    chunk_size = 14
    for chunk_idx in range(0, len(headers), chunk_size):
        chunk = headers[chunk_idx : chunk_idx + chunk_size]
        if chunk_idx > 0:
            story.append(Spacer(1, 8))
            cont = Table(
                [[Paragraph(f"<i>Continuación — columnas {chunk_idx + 1}–{chunk_idx + len(chunk)}</i>", hdr_style)]],
                colWidths=[page_size[0] - 12 * mm],
            )
            cont.setStyle(TableStyle([
                ("GRID", (0, 0), (-1, -1), 0.35, border),
                ("BACKGROUND", (0, 0), (-1, -1), sub_bg),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]))
            story.append(cont)
            story.append(Spacer(1, 4))

        col_w = (page_size[0] - 12 * mm) / max(len(chunk), 1)
        row_h1 = [Paragraph(_mud_pdf_escape(h1), hdr_style) for _, h1, _ in chunk]
        row_h2 = [Paragraph(_mud_pdf_escape(h2), hdr_style) for _, _, h2 in chunk]
        data_rows = [row_h1, row_h2]
        for _, row in view_df.iterrows():
            data_rows.append([
                Paragraph(_mud_pdf_escape(_mud_pdf_format_cell(row.get(col_name))), cell_style)
                for col_name, _, _ in chunk
            ])

        tbl = Table(data_rows, colWidths=[col_w] * len(chunk), repeatRows=2)
        tbl.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.25, border),
            ("BACKGROUND", (0, 0), (-1, 0), hdr_bg),
            ("BACKGROUND", (0, 1), (-1, 1), sub_bg),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
            ("ROWBACKGROUNDS", (0, 2), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
        ]))
        story.append(tbl)

    doc.build(story)
    return buffer.getvalue()


def _send_mud_bitacora_email(
    attachment_bytes: bytes,
    to_email: str,
    subject: str,
    body: str,
    filename: str = "mud_bitacora.xlsx",
    smtp_server: str = MUD_SMTP_SERVER,
    smtp_port: int = MUD_SMTP_PORT,
    smtp_user: str = MUD_SMTP_USER,
    smtp_pass: str = MUD_SMTP_PASS,
    from_email: str = MUD_SMTP_FROM,
) -> tuple[bool, str]:
    """Envía la bitácora Excel por correo como adjunto."""
    if not smtp_user or not smtp_pass:
        return False, "Faltan credenciales SMTP. Configura MUD_SMTP_USER y MUD_SMTP_PASS en secrets."
    try:
        msg = EmailMessage()
        msg["Subject"] = subject
        msg["From"] = from_email
        msg["To"] = to_email
        msg.set_content(body)
        msg.add_attachment(
            attachment_bytes,
            maintype="application",
            subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=filename,
        )
        with smtplib.SMTP(smtp_server, smtp_port, timeout=30) as server:
            server.starttls()
            server.login(smtp_user, smtp_pass)
            server.send_message(msg)
        return True, f"Bitácora enviada correctamente a {to_email}."
    except Exception as e:
        return False, str(e)


def render_mud_report(to_email: str = "") -> None:
    _ms = st.session_state.get("mud_data_source")
    if _ms == "Correo electrónico":
        st.session_state["mud_data_source"] = MUD_SRC_EMAIL
    elif _ms == "Subir archivos":
        st.session_state["mud_data_source"] = MUD_SRC_FILES

    title_badges = textwrap.dedent(
        """
        <div style="margin-bottom: 0.5rem;">
            <span style="font-size: 1.5rem; font-weight: 600;">Mud Report</span>
            <span style="display: inline-flex; align-items: center; gap: 0.35rem; margin-left: 0.75rem; flex-wrap: wrap;">
                <span style="background: linear-gradient(135deg, #b91c1c 0%, #ea580c 50%, #f59e0b 100%); color: #fff; font-size: 0.7rem; font-weight: 700; padding: 0.22rem 0.6rem; border-radius: 999px; letter-spacing: 0.03em; box-shadow: 0 1px 3px rgba(234,88,12,0.4);">🔥 Rogii</span>
                <span style="background: linear-gradient(135deg, #0f766e 0%, #14b8a6 100%); color: #fff; font-size: 0.7rem; font-weight: 600; padding: 0.2rem 0.55rem; border-radius: 999px;">Bitácora</span>
                <span style="background: linear-gradient(135deg, #1e3a5f 0%, #2563eb 100%); color: #fff; font-size: 0.7rem; font-weight: 600; padding: 0.2rem 0.55rem; border-radius: 999px;">PDF / Excel / CSV</span>
                <span style="background: linear-gradient(135deg, #7c2d12 0%, #ea580c 100%); color: #fff; font-size: 0.7rem; font-weight: 600; padding: 0.2rem 0.55rem; border-radius: 999px;">Correo</span>
            </span>
        </div>
        """
    )
    st.markdown(title_badges, unsafe_allow_html=True)
    st.caption('Carga reportes de lodo en PDF, Excel o CSV (subiendo archivos o desde correo). Se genera una bitácora unificada por día.')

    mud_source = st.radio(
        'Fuente de datos',
        [MUD_SRC_FILES, MUD_SRC_EMAIL],
        horizontal=True,
        key="mud_data_source",
        format_func=lambda x: 'Subir archivos' if x == MUD_SRC_FILES else 'Correo electrónico',
    )

    parsed: list[dict] = []

    # Chips de contexto (Rogii + fuente + Auto 60s si aplica)
    mud_chip_items = [
        ("🔥 Rogii", "#b91c1c", "#ea580c"),
        ('Correo electrónico', "#1e3a5f", "#2563eb")
        if mud_source == MUD_SRC_EMAIL
        else ('📁 Subir archivos', "#0f766e", "#14b8a6"),
    ]
    if mud_source == MUD_SRC_EMAIL and st.session_state.get("mud_auto_refresh", False):
        mud_chip_items.append(("Auto 60s 🔥", "#7c2d12", "#ea580c"))
    mud_cols = st.columns(len(mud_chip_items))
    for i, (label, c1, c2) in enumerate(mud_chip_items):
        with mud_cols[i]:
            st.markdown(
                f'<span style="display:inline-flex;align-items:center;gap:0.25rem;'
                f"background:linear-gradient(135deg,{c1},{c2});color:#fff;font-size:0.75rem;font-weight:600;"
                f'padding:0.25rem 0.6rem;border-radius:999px;box-shadow:0 1px 2px rgba(0,0,0,0.2);">{label}</span>',
                unsafe_allow_html=True,
            )
    st.markdown("<div style='height:6px'></div>", unsafe_allow_html=True)

    if mud_source == MUD_SRC_EMAIL:
        with st.expander('Configuración de correo (IMAP)', expanded=True):
            st.caption('Credenciales cargadas desde `.streamlit/secrets.toml` (IMAP_SERVER, IMAP_USER, IMAP_PASS). Puedes editarlas aquí solo para esta sesión.')
            col_imap1, col_imap2 = st.columns(2)
            with col_imap1:
                imap_server = st.text_input(
                    'Servidor IMAP',
                    value=MUD_IMAP_SERVER,
                    key="mud_imap_server",
                    help='Ej: imap.gmail.com',
                )
                imap_user = st.text_input(
                    'Usuario (correo)',
                    value=MUD_IMAP_USER,
                    key="mud_imap_user",
                )
            with col_imap2:
                imap_pass = st.text_input(
                    'Contraseña (App Password en Gmail)',
                    value=MUD_IMAP_PASS,
                    type="password",
                    key="mud_imap_pass",
                    help='En Gmail usa una contraseña de aplicación, no la de la cuenta.',
                )
                filename_filter = st.text_input(
                    'Filtrar por nombre de archivo (opcional)',
                    value=MUD_IMAP_FILTER,
                    placeholder='Ej: "Daily Full Report" o "LA-358"',
                    key="mud_imap_filter",
                )
            mark_read = st.checkbox(
                'Marcar correos como leídos al descargar',
                value=True,
                key="mud_imap_mark_read",
            )

        st.markdown('**Revisión automática**')
        mud_auto_refresh = st.checkbox(
            'Revisar correo automáticamente cada 60 s',
            value=st.session_state.get("mud_auto_refresh", False),
            key="mud_auto_refresh",
            help='Cada X segundos se consulta el correo y se actualiza la bitácora. Desmarca para detener.',
        )
        if mud_auto_refresh:
            mud_refresh_interval = st.number_input(
                'Intervalo (segundos)',
                min_value=30,
                max_value=300,
                value=60,
                step=15,
                key="mud_auto_refresh_interval",
                help='Cada cuántos segundos se revisa el correo (30–300 s).',
            )

        run_fetch = st.button(
            '🔥 Rogii – Revisar correo y cargar reportes',
            type="primary",
            key="mud_fetch_email_btn",
            help='Consulta IMAP y descarga adjuntos PDF/Excel/CSV de correos no leídos.',
        ) or (
            mud_auto_refresh
            and st.session_state.pop("mud_auto_rerun_trigger", False)
        )

        if run_fetch:
            if not imap_server or not imap_user or not imap_pass:
                st.error('Completa servidor IMAP, usuario y contraseña (o configúralos en .env).')
            else:
                with st.spinner('Conectando al correo y descargando adjuntos...'):
                    try:
                        attachments = _fetch_mud_attachments_from_email(
                            imap_server.strip(),
                            imap_user.strip(),
                            imap_pass.strip(),
                            filename_contains=filename_filter.strip() or None,
                            mark_read=mark_read,
                        )
                    except Exception as e:
                        st.error(f"{'No se pudo conectar o descargar:'} {e}")
                        attachments = []
                if not attachments:
                    st.info('No se encontraron adjuntos PDF/Excel/CSV en correos no leídos (o no coinciden con el filtro).')
                else:
                    st.success('Se descargaron **{n}** adjunto(s). Procesando...'.format(n=len(attachments)))
                    for name, data in attachments:
                        try:
                            parsed.extend(_mud_parse_attachment(name, data))
                        except Exception as e:
                            st.warning(f"No se pudo procesar **{name}**: {e}")
                    if parsed:
                        merged = _mud_merge_bitacora(
                            st.session_state.get("mud_bitacora"), _build_mud_bitacora(parsed)
                        )
                        st.session_state["mud_bitacora"] = merged
                        st.success(f"Bitácora actualizada con **{len(parsed)}** registro(s) desde correo.")
                        st.rerun()
                    else:
                        st.warning("No se detectaron propiedades de lodo en los adjuntos.")

    else:
        uploaded = st.file_uploader(
            'Subir reportes de lodo (PDF, Excel, CSV) — puedes soltar varios días a la vez',
            type=["pdf", "xlsx", "xls", "csv"],
            accept_multiple_files=True,
            key="mud_upload",
        )

        col_acc, col_clear = st.columns([3, 1])
        with col_acc:
            accumulate = st.toggle(
                'Acumular reportes en una sola bitácora',
                value=st.session_state.get("mud_accumulate", True),
                key="mud_accumulate",
                help='Activado: cada reporte que subas se suma a la bitácora, ordenada por fecha y hora, '
                     'y las muestras repetidas (misma fecha/hora, origen y pozo) se actualizan en vez de duplicarse. '
                     'Desactivado: la bitácora refleja solo los archivos que están ahora en el cargador.',
            )
        with col_clear:
            if st.button('🗑️ Limpiar bitácora', key="mud_clear_btn",
                         help='Vacía la bitácora acumulada. Los archivos del cargador se vuelven a leer.'):
                for k in ("mud_bitacora", "mud_ingested", "mud_parse_cache"):
                    st.session_state.pop(k, None)
                st.rerun()

        if uploaded:
            cache = st.session_state.setdefault("mud_parse_cache", {})
            ingested = st.session_state.setdefault("mud_ingested", {})
            current_sigs: list[str] = []
            new_sigs: list[str] = []
            for f in uploaded:
                name = getattr(f, "name", "") or ""
                try:
                    data = f.getvalue()
                except Exception:
                    f.seek(0)
                    data = f.read()
                sig = _mud_file_signature(data)
                current_sigs.append(sig)
                # El parseo se cachea por contenido: Streamlit reejecuta el script en cada
                # interacción y el cargador conserva los archivos, así que sin esto cada
                # clic volvería a leer todos los PDFs.
                if sig not in cache:
                    try:
                        cache[sig] = _mud_parse_attachment(name, data)
                    except Exception as e:
                        cache[sig] = []
                        st.warning(f"No se pudo procesar **{name}**: {e}")
                    if not cache[sig]:
                        st.warning(f"No se detectaron propiedades de lodo en **{name}**.")
                if sig not in ingested:
                    new_sigs.append(sig)
                    ingested[sig] = name

            if accumulate:
                if new_sigs:
                    fresh = [r for s in new_sigs for r in cache.get(s, [])]
                    if fresh:
                        st.session_state["mud_bitacora"] = _mud_merge_bitacora(
                            st.session_state.get("mud_bitacora"), _build_mud_bitacora(fresh)
                        )
            else:
                rows = [r for s in current_sigs for r in cache.get(s, [])]
                st.session_state["mud_bitacora"] = (
                    _mud_merge_bitacora(None, _build_mud_bitacora(rows)) if rows else pd.DataFrame()
                )

    bitacora = st.session_state.get("mud_bitacora")
    if bitacora is None or bitacora.empty:
        st.info("Sube uno o más reportes (PDF, Excel o CSV) para generar la bitácora.")
        return
    bitacora_view = _mud_build_view_df(bitacora)

    # Chips pro Rogii sobre la bitácora
    n_reg = len(bitacora)
    mud_days = pd.to_datetime(bitacora["Date"], errors="coerce").dt.normalize().dropna()
    n_days = int(mud_days.nunique())
    mud_wells = sorted(
        w for w in bitacora.get("Well", pd.Series(dtype=str)).astype(str).str.strip().unique() if w and w != "nan"
    )
    bitacora_chips = [
        ("🔥 Rogii", "#b91c1c", "#ea580c"),
        (f"{n_reg:,} muestras", "#0f766e", "#14b8a6"),
        (f"{n_days} día{'s' if n_days != 1 else ''}", "#1e3a5f", "#2563eb"),
    ]
    if bitacora["Date"].notna().any():
        d_min = bitacora["Date"].min()
        d_max = bitacora["Date"].max()
        if hasattr(d_min, "strftime"):
            bitacora_chips.append((f"{d_min.strftime('%d/%m')} – {d_max.strftime('%d/%m')}", "#334155", "#64748b"))
    if len(mud_wells) == 1:
        bitacora_chips.append((mud_wells[0][:26], "#3f3f46", "#71717a"))
    bitacora_cols = st.columns(len(bitacora_chips))
    for i, (label, c1, c2) in enumerate(bitacora_chips):
        with bitacora_cols[i]:
            st.markdown(
                f'<span style="display:inline-flex;align-items:center;gap:0.25rem;'
                f"background:linear-gradient(135deg,{c1},{c2});color:#fff;font-size:0.75rem;font-weight:600;"
                f'padding:0.28rem 0.65rem;border-radius:999px;box-shadow:0 1px 3px rgba(0,0,0,0.15);">{label}</span>',
                unsafe_allow_html=True,
            )
    st.success(
        f"Bitácora: **{n_reg:,}** muestras en **{n_days}** día(s), ordenadas por fecha y hora. "
        "«Properties» reinicia en 1 cada jornada."
    )
    if len(mud_wells) > 1:
        st.warning(
            "La bitácora mezcla **{n}** pozos ({lista}). Se conservan por separado, pero el formato "
            "exportado no tiene columna de pozo: si esperabas uno solo, limpia la bitácora y vuelve a cargar.".format(
                n=len(mud_wells), lista=", ".join(mud_wells[:4]) + ("…" if len(mud_wells) > 4 else "")
            )
        )
    # Un solo filtro arriba, que escopa gráficas y estadísticas por igual. No toca la
    # bitácora ni las descargas: esas siempre salen con todo lo acumulado.
    df_charts = bitacora
    if n_days > 1:
        d_lo, d_hi = mud_days.min().date(), mud_days.max().date()
        sel_lo, sel_hi = st.slider(
            "Ventana de análisis",
            min_value=d_lo, max_value=d_hi, value=(d_lo, d_hi),
            format="DD/MM/YY", key="mud_window",
            help="Recorta las gráficas y las estadísticas a un rango de días. La bitácora y las descargas no se filtran.",
        )
        _dates_only = pd.to_datetime(bitacora["Date"], errors="coerce").dt.date
        df_charts = bitacora[(_dates_only >= sel_lo) & (_dates_only <= sel_hi)]
        if len(df_charts) != n_reg:
            st.caption(f"Ventana activa: **{len(df_charts):,}** de {n_reg:,} muestras.")

    tab_bitacora, tab_graficas, tab_stats = st.tabs(["Bitácora", "Gráficas y evolución", "Estadísticas"])

    with tab_bitacora:
        st.subheader("Bitácora de propiedades de fluidos")
        mud_lang = st.radio(
            "Encabezados",
            [MUD_LANG_ES, MUD_LANG_EN],
            horizontal=True,
            key="mud_header_lang",
            format_func=lambda l: "Español (etiquetas del reporte)" if l == MUD_LANG_ES else "Inglés (formato WellSight)",
            help="En español los encabezados son los del propio reporte de lodo (Aceite, PC, R600, Exc. Cal…). "
                 "En inglés son los del formato WellSight original (NAP, YP, Lectura 600, Excess Lime…).",
        )
        # Lo que se ve en pantalla es lo que sale en el archivo.
        mud_specs = _mud_localize_specs(_mud_export_specs_for(bitacora_view), mud_lang)
        mud_rename = {col: h1 for col, h1, _ in mud_specs}
        st.dataframe(
            _mud_display_df(bitacora_view).rename(columns=mud_rename),
            use_container_width=True,
            hide_index=True,
        )

        with st.expander("¿De dónde sale cada columna?", expanded=False):
            st.caption("Equivalencia entre la columna de la bitácora y la etiqueta del reporte de lodo.")
            # En markdown y no en st.dataframe: la tabla es para leerse, así el texto
            # ajusta, se puede buscar con Ctrl+F y no queda recortado en un grid.
            st.markdown(
                "| Columna de la bitácora | Etiqueta en el reporte | Nota |\n|---|---|---|\n"
                + "\n".join(f"| {c} | {o} | {n} |" for c, o, n in MUD_COLUMN_GLOSSARY)
            )

        buf_csv = io.BytesIO()
        bitacora_view.rename(columns=mud_rename).to_csv(buf_csv, index=False, encoding="utf-8-sig")
        buf_csv.seek(0)
        xlsx_bytes = _export_mud_bitacora_excel(bitacora_view, lang=mud_lang)
        pdf_bytes = _export_mud_bitacora_pdf(bitacora_view, lang=mud_lang)

        default_base = _default_mud_bitacora_basename(bitacora)
        st.markdown("### 📎 Nombre del archivo de salida")
        output_base = st.text_input(
            "Nombre base para CSV, Excel, PDF y adjunto de correo:",
            value=default_base,
            key="mud_output_basename",
            help="Puedes editarlo antes de descargar o enviar. Se añade .csv, .xlsx o .pdf según el formato.",
        )
        output_base = _sanitize_filename(
            output_base.replace(".csv", "").replace(".xlsx", "").replace(".xls", "").replace(".pdf", "")
        )
        csv_name = f"{output_base}.csv"
        xlsx_name = f"{output_base}.xlsx"
        pdf_name = f"{output_base}.pdf"
        st.caption(
            f"Descarga: **{csv_name}** · **{xlsx_name}** · **{pdf_name}** · Correo adjunta: **{xlsx_name}**"
        )

        # Mismo destinatario que "To email parsing" en el sidebar (compartido con Daily
        # Report): un solo campo controla a dónde se envían ambos reportes.
        mud_to_email = to_email.strip() or MUD_SMTP_TO
        st.caption(f"Se enviará a: **{mud_to_email}** (editable en «To email parsing», en el panel lateral).")

        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.download_button(
                "Exportar bitácora (CSV)",
                data=buf_csv.getvalue(),
                file_name=csv_name,
                mime="text/csv",
                key="mud_export_csv",
            )
        with col2:
            st.download_button(
                "Exportar bitácora (Excel)",
                data=xlsx_bytes,
                file_name=xlsx_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="mud_export_xlsx",
            )
        with col3:
            st.download_button(
                "Exportar bitácora (PDF)",
                data=pdf_bytes,
                file_name=pdf_name,
                mime="application/pdf",
                key="mud_export_pdf",
            )
        with col4:
            if st.button("Enviar bitácora por correo", key="mud_send_email_btn", type="secondary"):
                if not mud_to_email.strip() or "@" not in mud_to_email:
                    st.error("Completa un correo destino válido antes de enviar.")
                else:
                    date_label = ""
                    try:
                        if "Date" in bitacora.columns and bitacora["Date"].notna().any():
                            dmax = pd.to_datetime(bitacora["Date"], errors="coerce").dropna().max()
                            if pd.notna(dmax):
                                date_label = dmax.strftime("%Y-%m-%d")
                    except Exception:
                        date_label = ""
                    subject = f"Mud bitácora {date_label}".strip()
                    body = (
                        "Hola,\n\n"
                        "Adjunto la bitácora de propiedades de fluidos generada desde la app.\n\n"
                        "Saludos."
                    )
                    ok, msg = _send_mud_bitacora_email(
                        attachment_bytes=xlsx_bytes,
                        to_email=mud_to_email.strip(),
                        subject=subject,
                        body=body,
                        filename=xlsx_name,
                    )
                    if ok:
                        st.success(msg)
                    else:
                        st.error(f"No se pudo enviar la bitácora por correo: {msg}")

        with st.expander("Configuración de envío por correo", expanded=False):
            st.caption("Servidor/usuario/clave se leen desde `.streamlit/secrets.toml` (SMTP_*). El destinatario es «To email parsing», en el panel lateral.")
            e1, e2 = st.columns(2)
            with e1:
                st.text_input("SMTP server", value=MUD_SMTP_SERVER, disabled=True, key="mud_smtp_server_view")
                st.text_input("SMTP user", value=MUD_SMTP_USER, disabled=True, key="mud_smtp_user_view")
                st.text_input("From", value=MUD_SMTP_FROM, disabled=True, key="mud_smtp_from_view")
            with e2:
                st.text_input("SMTP port", value=str(MUD_SMTP_PORT), disabled=True, key="mud_smtp_port_view")
                st.text_input("SMTP password", value=("********" if MUD_SMTP_PASS else ""), type="password", disabled=True, key="mud_smtp_pass_view")

    with tab_graficas:
        st.subheader("Evolución de propiedades")
        props = _mud_numeric_property_columns(bitacora)
        if not props:
            st.info("No hay columnas numéricas para graficar.")
        else:
            st.caption(
                "Todas las gráficas se descargan en PNG de alta definición desde el ícono 📷 "
                "de la barra de herramientas (5400 × 3000 px)."
            )
            evo_mode = st.radio(
                "Vista",
                [_MUD_EVO_PANELS, _MUD_EVO_NORM, _MUD_EVO_RAW],
                horizontal=True,
                key="mud_evo_mode",
                format_func=lambda m: {
                    _MUD_EVO_PANELS: "Paneles (un eje por propiedad)",
                    _MUD_EVO_NORM: "Superpuesto normalizado 0–100",
                    _MUD_EVO_RAW: "Superpuesto en valor real",
                }[m],
                help="Paneles: cada propiedad con su propia escala, recomendado para varios días. "
                     "Normalizado: compara formas entre magnitudes distintas. "
                     "Valor real: útil solo entre propiedades de escala parecida.",
            )

            with st.expander("Propiedades a graficar", expanded=True):
                # Marcadas por defecto las que de verdad se siguen día a día, no las
                # primeras cuatro de la lista (que caían en temperaturas de medición).
                defaults = [p for p in MUD_DEFAULT_CHART_PROPS if p in props] or props[:4]
                n_cols = 4
                n_rows = (len(props) + n_cols - 1) // n_cols
                checkbox_state = {}
                for row in range(n_rows):
                    cols = st.columns(n_cols)
                    for col_idx in range(n_cols):
                        i = row * n_cols + col_idx
                        if i >= len(props):
                            break
                        p = props[i]
                        default = p in defaults
                        checkbox_state[p] = st.checkbox(
                            p,
                            value=st.session_state.get(f"mud_cb_{p}", default),
                            key=f"mud_cb_{p}",
                            label_visibility="visible",
                        )
            selected = [p for p in props if checkbox_state.get(p, False)]
            selected = [p for p in selected if pd.to_numeric(df_charts.get(p), errors="coerce").notna().any()]

            if not selected:
                st.info("Marca al menos una propiedad con datos en la ventana seleccionada.")
            else:
                limit = _MUD_MAX_PANELS if evo_mode == _MUD_EVO_PANELS else MUD_MAX_OVERLAY_SERIES
                if len(selected) > limit:
                    st.info(
                        f"Se muestran las primeras **{limit}** propiedades de las {len(selected)} marcadas: "
                        "más allá de ese número la gráfica deja de ser legible."
                    )
                    selected = selected[:limit]

                if evo_mode == _MUD_EVO_PANELS:
                    fig = _mud_evolution_panels_figure(df_charts, selected)
                else:
                    if evo_mode == _MUD_EVO_RAW and len(selected) > 1:
                        decades = []
                        for p in selected:
                            s = pd.to_numeric(df_charts[p], errors="coerce").dropna().abs()
                            s = s[s > 0]
                            if len(s):
                                decades.append(float(np.log10(s.median())))
                        if decades and (max(decades) - min(decades)) >= 1.5:
                            st.warning(
                                "Las propiedades marcadas están en escalas muy distintas (p. ej. densidad ~1680 kg/m³ "
                                "frente a YP ~10 lb/100ft²): en un solo eje las pequeñas quedan planas. "
                                "Usa **Paneles** o **Superpuesto normalizado**."
                            )
                    fig = _mud_evolution_overlay_figure(
                        df_charts, selected, _mud_assign_series_slots(selected),
                        normalize=(evo_mode == _MUD_EVO_NORM),
                    )
                st.plotly_chart(
                    fig, use_container_width=True,
                    config=_mud_hd_config(f"{_default_mud_bitacora_basename(bitacora)}_evolucion"),
                )
                resumen = " · ".join(
                    f"**{p}** {format_num(pd.to_numeric(df_charts[p], errors='coerce').min())}–"
                    f"{format_num(pd.to_numeric(df_charts[p], errors='coerce').max())}"
                    for p in selected[:6]
                )
                if resumen:
                    st.caption(f"Rango en la ventana — {resumen}")

            st.markdown("---")
            st.subheader("Detalle por propiedad")
            single_prop = st.selectbox("Propiedad", ["(ninguna)"] + props, key="mud_single_prop")
            if single_prop and single_prop != "(ninguna)":
                s = pd.to_numeric(df_charts[single_prop], errors="coerce").dropna()
                if len(s):
                    st.plotly_chart(
                        _mud_property_detail_figure(df_charts, single_prop, _mud_palette()[0]),
                        use_container_width=True,
                        config=_mud_hd_config(f"{_sanitize_filename(single_prop)}_evolucion"),
                    )
                    _render_chips_row([
                        (f"n={len(s):,}", "gray"),
                        (f"min {format_num(s.min())}", "gray"),
                        (f"max {format_num(s.max())}", "gray"),
                        (f"promedio {format_num(s.mean())}", "green"),
                    ])
                else:
                    st.info("Esa propiedad no tiene datos en la ventana seleccionada.")

            st.markdown("---")
            st.subheader("Correlación, control y perfil")

            # 1) Heatmap de correlación entre propiedades de lodo
            corr_props = [p for p in props if pd.to_numeric(df_charts.get(p), errors="coerce").nunique() > 1]
            if len(corr_props) >= 2 and len(df_charts) >= 3:
                st.markdown("**Correlación entre propiedades**")
                corr_mud = df_charts[corr_props].corr(numeric_only=True)
                if not corr_mud.isna().all().all():
                    corr_pct = (corr_mud * 100).round(0)
                    text_arr = np.where(
                        np.isnan(corr_pct.values),
                        "",
                        (np.nan_to_num(corr_pct.values, nan=0.0).astype(int)).astype(str) + "%",
                    )
                    _t = _mud_viz_tokens()
                    fig_corr_mud = px.imshow(
                        corr_mud,
                        color_continuous_scale=_mud_correlation_colorscale(),
                        zmin=-1,
                        zmax=1,
                    )
                    fig_corr_mud.update_traces(
                        text=text_arr,
                        texttemplate="%{text}",
                        textfont=dict(size=10),
                        xgap=2,
                        ygap=2,
                        hovertemplate="%{y} ↔ %{x}: %{z:.2f}<extra></extra>",
                    )
                    fig_corr_mud.update_layout(
                        coloraxis_colorbar=dict(
                            title=dict(text="Correlación", font=dict(size=11, color=_t["secondary"])),
                            tickvals=[-1, -0.5, 0, 0.5, 1], tickfont=dict(size=10, color=_t["muted"]),
                            outlinewidth=0, thickness=12, len=0.8,
                        )
                    )
                    fig_corr_mud = _mud_hd_theme(
                        fig_corr_mud,
                        h=max(420, 22 * len(corr_props) + 140),
                        title="Correlación lineal entre propiedades",
                        legend=False, hovermode="closest", spikes=False,
                    )
                    fig_corr_mud.update_xaxes(showgrid=False, tickangle=-45, showline=False)
                    fig_corr_mud.update_yaxes(showgrid=False)
                    fig_corr_mud.update_layout(margin=dict(l=150, r=26, t=62, b=140))
                    st.plotly_chart(
                        fig_corr_mud, use_container_width=True,
                        config=_mud_hd_config(f"{_default_mud_bitacora_basename(bitacora)}_correlacion"),
                    )
                    st.caption(
                        "Rojo = suben juntas, azul = una sube cuando la otra baja, gris = sin relación lineal. "
                        "Los valores exactos están en la celda; la matriz completa se puede descargar en HD."
                    )
                else:
                    st.info("No hay suficientes datos para calcular correlaciones.")
            else:
                st.caption("Se necesitan al menos 3 muestras y 2 propiedades con variación para la correlación.")

            # 2) Gráfico de control (propiedad vs fecha, media ± 2σ)
            st.markdown("**Gráfico de control (media ± 2σ)**")
            ctrl_prop = st.selectbox(
                "Propiedad para control",
                props,
                key="mud_ctrl_prop",
                help="Los puntos fuera de la banda se marcan con rombo y color de alerta.",
            )
            if ctrl_prop:
                df_ctrl = df_charts[["Date", ctrl_prop]].dropna().sort_values("Date")
                if len(df_ctrl) >= 2:
                    _t = _mud_viz_tokens()
                    base = _mud_palette()[0]
                    mean_val = float(df_ctrl[ctrl_prop].mean())
                    std_val = float(df_ctrl[ctrl_prop].std()) or 1e-6
                    upper = mean_val + 2 * std_val
                    lower = mean_val - 2 * std_val
                    df_ctrl = df_ctrl.copy()
                    df_ctrl["_out"] = (df_ctrl[ctrl_prop] > upper) | (df_ctrl[ctrl_prop] < lower)
                    in_spec = df_ctrl[~df_ctrl["_out"]]
                    out_spec = df_ctrl[df_ctrl["_out"]]
                    fig_ctrl = go.Figure()
                    fig_ctrl.add_trace(go.Scatter(
                        x=list(df_ctrl["Date"]) + list(df_ctrl["Date"])[::-1],
                        y=[upper] * len(df_ctrl) + [lower] * len(df_ctrl),
                        fill="toself", fillcolor=_mud_rgba(base, 0.08), line=dict(width=0),
                        hoverinfo="skip", name="Banda ±2σ",
                    ))
                    fig_ctrl.add_trace(go.Scatter(
                        x=df_ctrl["Date"], y=df_ctrl[ctrl_prop], mode="lines",
                        line=dict(width=1.5, color=_mud_rgba(base, 0.55)),
                        hoverinfo="skip", showlegend=False,
                    ))
                    if not in_spec.empty:
                        fig_ctrl.add_trace(go.Scatter(
                            x=in_spec["Date"], y=in_spec[ctrl_prop], mode="markers",
                            name="Dentro de límites",
                            marker=dict(size=9, color=base, line=dict(width=2, color=_t["ring"])),
                            hovertemplate=f"<b>{ctrl_prop}</b> %{{y}}<extra></extra>",
                        ))
                    if not out_spec.empty:
                        # Color de estado + símbolo + etiqueta en la leyenda: la alerta nunca
                        # se apoya solo en el color.
                        fig_ctrl.add_trace(go.Scatter(
                            x=out_spec["Date"], y=out_spec[ctrl_prop], mode="markers",
                            name="⚠ Fuera de ±2σ",
                            marker=dict(size=13, color=_MUD_STATUS["critical"], symbol="diamond",
                                        line=dict(width=2, color=_t["ring"])),
                            hovertemplate=f"<b>{ctrl_prop}</b> %{{y}} — fuera de límites<extra></extra>",
                        ))
                    for y_val, lbl in ((mean_val, "Media"), (upper, "UCL"), (lower, "LCL")):
                        fig_ctrl.add_hline(
                            y=y_val, line=dict(color=_t["secondary"], width=1, dash="dash"),
                            annotation_text=f"{lbl} {format_num(y_val)}",
                            annotation_position="right",
                            annotation_font=dict(size=10, color=_t["secondary"]),
                        )
                    fig_ctrl = _mud_hd_theme(fig_ctrl, h=430, title=f"Control — {ctrl_prop}",
                                             y_title=ctrl_prop)
                    fig_ctrl.update_xaxes(title_text="Fecha / hora")
                    fig_ctrl.update_layout(margin=dict(r=96))
                    _mud_style_time_axis(fig_ctrl, df_ctrl["Date"])
                    st.plotly_chart(
                        fig_ctrl, use_container_width=True,
                        config=_mud_hd_config(f"{_sanitize_filename(ctrl_prop)}_control"),
                    )
                    st.caption(f"Media = {format_num(mean_val)}, LCL = {format_num(lower)}, UCL = {format_num(upper)}. Puntos fuera de banda = {len(out_spec)}.")
                else:
                    st.info("Se necesitan al menos 2 puntos para el gráfico de control.")

            # 3) Radar / perfil del lodo: compara dos muestras cualesquiera
            st.markdown("**Perfil radar (comparación normalizada)**")
            radar_props = [p for p in props if pd.to_numeric(df_charts.get(p), errors="coerce").notna().any()]
            if len(radar_props) > MUD_MAX_OVERLAY_SERIES + 2:
                radar_props = radar_props[:MUD_MAX_OVERLAY_SERIES + 2]
            dates_opt = sorted(df_charts["Date"].dropna().unique(), reverse=True)
            if len(radar_props) >= 3 and len(dates_opt) > 0:
                def _fmt_dt(x):
                    return pd.Timestamp(x).strftime("%d/%m %H:%M")

                col_r1, col_r2 = st.columns(2)
                with col_r1:
                    radar_date = st.selectbox("Muestra", options=dates_opt, format_func=_fmt_dt,
                                              index=0, key="mud_radar_date")
                with col_r2:
                    cmp_opts = ["(sin comparar)"] + [d for d in dates_opt if d != radar_date]
                    radar_cmp = st.selectbox("Comparar contra", options=cmp_opts,
                                             format_func=lambda x: x if isinstance(x, str) else _fmt_dt(x),
                                             index=0, key="mud_radar_cmp")

                def _radar_values(when):
                    rows_at = df_charts[df_charts["Date"] == when]
                    if rows_at.empty:
                        return None, None
                    row = rows_at.iloc[-1]
                    out, raw = [], []
                    for p in radar_props:
                        v = pd.to_numeric(pd.Series([row.get(p)]), errors="coerce").iloc[0]
                        col = pd.to_numeric(df_charts[p], errors="coerce").dropna()
                        if pd.isna(v) or col.empty or col.max() == col.min():
                            out.append(50.0 if not pd.isna(v) else 0.0)
                        else:
                            out.append(round((float(v) - float(col.min())) / (float(col.max()) - float(col.min())) * 100, 1))
                        raw.append(v)
                    return out, raw

                pal = _mud_palette()
                fig_radar = go.Figure()
                series = [(radar_date, pal[0])]
                if not isinstance(radar_cmp, str):
                    series.append((radar_cmp, pal[1]))
                for when, color in series:
                    vals, raw = _radar_values(when)
                    if vals is None:
                        continue
                    fig_radar.add_trace(go.Scatterpolar(
                        r=vals + [vals[0]],
                        theta=radar_props + [radar_props[0]],
                        customdata=list(raw) + [raw[0]],
                        fill="toself", name=_fmt_dt(when),
                        line=dict(color=color, width=2),
                        fillcolor=_mud_rgba(color, 0.18),
                        marker=dict(size=7, color=color),
                        hovertemplate="%{theta}: %{customdata} (índice %{r})<extra>" + _fmt_dt(when) + "</extra>",
                    ))
                _t = _mud_viz_tokens()
                fig_radar = _mud_hd_theme(fig_radar, h=520, title="Perfil normalizado del lodo",
                                          legend=len(fig_radar.data) > 1, hovermode="closest", spikes=False)
                fig_radar.update_layout(polar=dict(
                    bgcolor="rgba(0,0,0,0)",
                    radialaxis=dict(visible=True, range=[0, 100], gridcolor=_t["grid"],
                                    linecolor=_t["axis"], tickfont=dict(size=10, color=_t["muted"])),
                    angularaxis=dict(gridcolor=_t["grid"], linecolor=_t["axis"],
                                     tickfont=dict(size=11, color=_t["secondary"])),
                ))
                st.plotly_chart(
                    fig_radar, use_container_width=True,
                    config=_mud_hd_config(f"{_default_mud_bitacora_basename(bitacora)}_perfil", width=1400, height=1400),
                )
                st.caption(
                    "Cada eje es la propiedad normalizada 0–100 % contra su min/max en la ventana. "
                    "El tooltip muestra el valor real. Comparar dos muestras deja ver de un golpe qué se movió entre jornadas."
                )
            else:
                st.caption("Se necesitan al menos 3 propiedades con datos para el perfil radar.")

    with tab_stats:
        st.subheader("Estadísticas por propiedad")
        props = _mud_numeric_property_columns(bitacora)
        if not props:
            st.info("No hay columnas numéricas.")
        else:
            stats_rows = []
            for p in props:
                s = pd.to_numeric(df_charts[p], errors="coerce").dropna()
                if not len(s):
                    continue
                mean_v = float(s.mean())
                std_v = float(s.std()) if len(s) > 1 else 0.0
                stats_rows.append({
                    "Propiedad": p,
                    "N": int(len(s)),
                    "Min": float(s.min()),
                    "P25": float(s.quantile(0.25)),
                    "Mediana": float(s.median()),
                    "Media": mean_v,
                    "P75": float(s.quantile(0.75)),
                    "Max": float(s.max()),
                    "Desv. est.": std_v,
                    # Coeficiente de variación: la manera de comparar estabilidad entre
                    # propiedades con unidades distintas.
                    "CV %": (std_v / abs(mean_v) * 100.0) if mean_v else np.nan,
                })
            if stats_rows:
                stats_df = pd.DataFrame(stats_rows)
                st.dataframe(
                    stats_df,
                    use_container_width=True,
                    hide_index=True,
                    column_config={
                        c: st.column_config.NumberColumn(c, format="%.2f")
                        for c in ("Min", "P25", "Mediana", "Media", "P75", "Max", "Desv. est.", "CV %")
                    },
                )
                st.download_button(
                    "⬇️ Descargar estadísticas (CSV)",
                    data=stats_df.to_csv(index=False).encode("utf-8-sig"),
                    file_name=f"{_default_mud_bitacora_basename(bitacora)}_estadisticas.csv",
                    mime="text/csv",
                    key="mud_stats_csv",
                )

                st.markdown("---")
                st.subheader("Qué se movió y qué se mantuvo estable")
                st.caption(
                    "Coeficiente de variación (σ / media) por propiedad, ordenado de mayor a menor. "
                    "Es la forma de comparar cuánto varió cada propiedad aunque estén en unidades distintas: "
                    "arriba lo que más se movió en la ventana, abajo lo que se mantuvo plano."
                )
                var_props = [r["Propiedad"] for r in stats_rows][:24]
                fig_var = _mud_variability_figure(df_charts, var_props)
                if fig_var.data:
                    st.plotly_chart(
                        fig_var, use_container_width=True,
                        config=_mud_hd_config(f"{_default_mud_bitacora_basename(bitacora)}_variabilidad",
                                              width=1600, height=1400),
                    )
                else:
                    st.info("Se necesitan al menos 2 muestras por propiedad para calcular la variabilidad.")

                st.markdown("---")
                st.subheader("Distribución (histograma)")
                prop_hist = st.selectbox("Propiedad para histograma", props, key="mud_hist_prop")
                if prop_hist:
                    vals = pd.to_numeric(df_charts[prop_hist], errors="coerce").dropna()
                    if len(vals):
                        st.plotly_chart(
                            _mud_distribution_figure(vals, prop_hist, _mud_palette()[0]),
                            use_container_width=True,
                            config=_mud_hd_config(f"{_sanitize_filename(prop_hist)}_distribucion"),
                        )
                        st.caption(f"**Resumen:** {series_summary(vals)}.")
                    else:
                        st.info("Esa propiedad no tiene datos en la ventana seleccionada.")

    # Auto-refresh correo cada N segundos (solo si fuente = Correo y hay bitácora)
    if (
        st.session_state.get("mud_data_source") == MUD_SRC_EMAIL
        and st.session_state.get("mud_auto_refresh")
    ):
        interval = int(st.session_state.get("mud_auto_refresh_interval", 60))
        interval = max(30, min(300, interval))
        countdown_placeholder = st.empty()
        for i in range(interval, 0, -1):
            countdown_placeholder.info('🔥 **Rogii** – Próxima revisión de correo en **{i}** s… (desmarca «Revisar correo automáticamente» para detener)'.format(i=i))
            time.sleep(1)
        countdown_placeholder.empty()
        st.session_state["mud_auto_rerun_trigger"] = True
        st.rerun()


if __name__ == "__main__":
    render_mud_report()
